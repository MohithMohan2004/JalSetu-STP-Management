from flask import Flask, jsonify, request, render_template, redirect, url_for
from flask import session
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

import threading
import uuid
from datetime import datetime, date, timedelta
import json
import os
import math
import requests
import pandas as pd
import csv
import osmnx as ox
import networkx as nx
from ml.predict_demand import predict_next_day, predict_week


def format_clean_address(address, lat, lon):
    try:
        place = (
            address.get("building") or
            address.get("amenity") or
            address.get("residential") or
            address.get("village") or
            address.get("hamlet")
        )
        area = address.get("suburb") or address.get("neighbourhood")
        road = address.get("road")
        district = address.get("state_district")
        state = address.get("state")
        pincode = address.get("postcode")
        formatted = ", ".join(filter(None, [place, area, road, district, state, pincode]))
        return formatted if formatted else f"{lat}, {lon}"
    except Exception as e:
        print("Format error:", e)
        return f"{lat}, {lon}"

app = Flask(__name__)

app.secret_key = "secret123"

# =========================================================
# LOAD ROAD NETWORK FOR A* ROUTING
# =========================================================
GRAPH_FILE = "bangalore_graph.graphml"

G = None

try:
    if os.path.exists(GRAPH_FILE):
        print("Loading saved road network...")
        G = ox.load_graphml(GRAPH_FILE)

        import random

        for u, v, k, data in G.edges(keys=True, data=True):
            traffic_factor = random.uniform(1.0, 3.0)

            data["traffic_factor"] = traffic_factor
            data["travel_cost"] = data["length"] * traffic_factor
    else:
        print("Skipping graph load (deployment)")
except Exception as e:
    print("Graph load skipped:", e)

print("Road network ready")

# =========================================================
# FILE PATHS
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

STP_FILE = os.path.join(BASE_DIR, "data", "stp_data.json")
STATUS_FILE = os.path.join(BASE_DIR, "data", "stp_status.json")
DATABASE_DIR = os.path.join(BASE_DIR, "database")
if not os.path.exists(DATABASE_DIR):
    os.makedirs(DATABASE_DIR)

# ✅ KEEP orders.csv INSIDE database/
ORDERS_FILE = os.path.join(DATABASE_DIR, "orders.csv")

PRICING_FILE = os.path.join(
    BASE_DIR,
    "data",
    "stp_pricing.csv"
)

STP_TRANSFERS_FILE = os.path.join(
    DATABASE_DIR,
    "stp_transfers.csv"
)

# =========================================================

TANKER_REGISTRATIONS_FILE = os.path.join(
    DATABASE_DIR,
    "tanker_registrations.csv"
)

STP_REGISTRATIONS_FILE = os.path.join(
    DATABASE_DIR,
    "stp_registrations.csv"
)
# =========================================================
# USER ACCOUNT DATABASE
# =========================================================

USERS_FILE = os.path.join(
    DATABASE_DIR,
    "users.xlsx"
)

users_lock = threading.Lock()

USER_FIELDS = [
    "user_id",
    "first_name",
    "last_name",
    "username",
    "mobile",
    "email",
    "password_hash",
    "role",
    "stp_id",
    "tanker_operator_id",
    "created_at",
    "account_status"
]

def ensure_users_file():
    """Create or safely update the Excel user database schema."""
    if not os.path.exists(USERS_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(USER_FIELDS)
        workbook.save(USERS_FILE)
        return

    with users_lock:
        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        existing_headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

        changed = False
        for field in USER_FIELDS:
            if field not in existing_headers:
                sheet.cell(row=1, column=sheet.max_column + 1, value=field)
                existing_headers.append(field)
                changed = True

        if changed:
            workbook.save(USERS_FILE)

        workbook.close()

def load_users():
    """Load all registered users from users.xlsx."""
    ensure_users_file()

    with users_lock:
        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]

        users = []
        for values in rows[1:]:
            user = {}
            for index, header in enumerate(headers):
                user[header] = values[index] if index < len(values) else ""
            users.append(user)

        return users

def append_user(user):
    """Append one user safely to users.xlsx."""
    ensure_users_file()

    with users_lock:
        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        # Ensure the expected header exists.
        existing_headers = [
            cell.value for cell in sheet[1]
        ]

        if existing_headers != USER_FIELDS:
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(USER_FIELDS)

        sheet.append([
            user.get(field, "") for field in USER_FIELDS
        ])

        workbook.save(USERS_FILE)


def safe_user_value(user, field_name, default=""):
    """Return a consistent string for legacy and newly migrated users."""
    value = user.get(field_name, default)
    if value is None:
        return ""
    return str(value).strip()


ensure_users_file()

# =========================================================

# SYNTHETIC / DEMAND HEATMAP DATASET
# =========================================================

DEMAND_CSV_FILE = os.path.join(
    DATABASE_DIR,
    "synthetic_orders.csv"
)

# =========================================================

# ENSURE FILES EXIST
# =========================================================
if not os.path.exists(STATUS_FILE):
    with open(STATUS_FILE, "w") as f:
        json.dump({}, f)

if not os.path.exists(ORDERS_FILE):
    with open(ORDERS_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "order_id",
            "stp_id",
            "stp_name",
            "quantity_kld",
            "quality",
            "water_type",
            "distance_km",
            "location",
            "buyer_name",
            "buyer_phone",
            "status",
            "created_at"
        ])

ORDER_FIELDS = [
    "order_id",
    "stp_id",
    "stp_name",
    "quantity_kld",
    "quality",
    "water_type",
    "distance_km",
    "location",
    "buyer_user_id",
    "buyer_name",
    "buyer_phone",
    "status",
    "created_at",
    "payment_status",
    "accepted_at",
    "capacity_release_at",
    "capacity_released"
]

STP_TRANSFER_FIELDS = [
    "transfer_id",
    "source_stp_id",
    "source_stp_name",
    "destination_stp_id",
    "destination_stp_name",
    "quantity_kld",
    "quality",
    "water_type",
    "distance_km",
    "status",
    "requested_at",
    "accepted_at",
    "rejected_at",
    "tanker_status",
    "delivered_at"
]

def ensure_stp_transfers_file():
    """Create or update the STP-to-STP transfer request CSV schema."""

    # Create the file if it does not exist or is empty
    if (
        not os.path.exists(STP_TRANSFERS_FILE)
        or os.path.getsize(STP_TRANSFERS_FILE) == 0
    ):
        with open(
            STP_TRANSFERS_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as f:

            writer = csv.DictWriter(
                f,
                fieldnames=STP_TRANSFER_FIELDS
            )

            writer.writeheader()

        return

    # Read the existing file
    with open(
        STP_TRANSFERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)

        existing_fields = reader.fieldnames or []

        rows = list(reader)

    # Nothing to change if schema is already current
    if existing_fields == STP_TRANSFER_FIELDS:
        return

    # Preserve all existing transfer data
    with open(
        STP_TRANSFERS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=STP_TRANSFER_FIELDS
        )

        writer.writeheader()

        for row in rows:

            writer.writerow({
                field: row.get(field, "")
                for field in STP_TRANSFER_FIELDS
            })

def ensure_orders_schema():
    """Add buyer_user_id to older orders.csv files without deleting existing orders."""
    if not os.path.exists(ORDERS_FILE) or os.path.getsize(ORDERS_FILE) == 0:
        with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
            writer.writeheader()
        return

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        existing_fields = reader.fieldnames or []
        rows = list(reader)

    if existing_fields == ORDER_FIELDS:
        return

    # Preserve every existing field and add the new account identifier.
    merged_fields = list(existing_fields)
    if "buyer_user_id" not in merged_fields:
        insert_at = merged_fields.index("buyer_name") if "buyer_name" in merged_fields else len(merged_fields)
        merged_fields.insert(insert_at, "buyer_user_id")

    # Keep the new canonical order.
    for field in ORDER_FIELDS:
        if field not in merged_fields:
            merged_fields.append(field)

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()

        for row in rows:
            normalized = {field: row.get(field, "") for field in ORDER_FIELDS}
            writer.writerow(normalized)

ensure_orders_schema()
ensure_stp_transfers_file()

# =========================================================
# HELPER FUNCTIONS
# =========================================================

def load_stps():
    with open(STP_FILE) as f:
        data = json.load(f)
        return data.get("stps", [])

def load_stp_pricing():
    if not os.path.exists(PRICING_FILE):
        return []

    with open(
        PRICING_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as file:

        return list(csv.DictReader(file))
    
def save_stps(stps):

    with open(STP_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    data["stps"] = stps

    with open(STP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

def auto_reset_capacity():
    """Release STP capacity for accepted orders exactly 24 hours after acceptance."""
    now = datetime.now()
    stps = load_stps()
    stps_changed = False
    orders_changed = False

    if not os.path.exists(ORDERS_FILE):
        return

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        orders = list(reader)

    for row in orders:
        status = (row.get("status") or "").strip()
        if status not in {"Accepted", "Out for Delivery", "Delivered"}:
            continue

        release_at_raw = (row.get("capacity_release_at") or "").strip()
        if not release_at_raw:
            accepted_at_raw = (row.get("accepted_at") or "").strip() or (row.get("created_at") or "").strip()
            try:
                accepted_at = datetime.fromisoformat(accepted_at_raw)
                release_at = accepted_at + timedelta(hours=24)
                row["accepted_at"] = accepted_at.isoformat()
                row["capacity_release_at"] = release_at.isoformat()
                row["capacity_released"] = row.get("capacity_released") or "False"
                release_at_raw = release_at.isoformat()
                orders_changed = True
            except (TypeError, ValueError):
                continue

        if str(row.get("capacity_released", "")).strip().lower() == "true":
            continue

        try:
            release_at = datetime.fromisoformat(release_at_raw)
        except (TypeError, ValueError):
            continue

        if now < release_at:
            continue

        quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
        for stp in stps:
            if str(stp.get("stp_id")) == str(row.get("stp_id")):
                total_capacity = float(stp.get("total_capacity_mld") or 0)
                available_capacity = float(stp.get("available_capacity_mld", total_capacity) or 0)
                stp["available_capacity_mld"] = min(total_capacity, available_capacity + quantity_mld)
                stp["current_load_mld"] = max(0.0, float(stp.get("current_load_mld", 0) or 0) - quantity_mld)
                stps_changed = True
                row["capacity_released"] = "True"
                orders_changed = True
                break

    if stps_changed:
        save_stps(stps)
    if orders_changed:
        with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
            writer.writeheader()
            for row in orders:
                writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})


def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) *
         math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c    
# =========================================================
# A* DISTANCE FUNCTION
# =========================================================
def astar_distance(lat1, lon1, lat2, lon2):
    
    if G is None:
        print("Using fallback distance")
        return haversine(lat1, lon1, lat2, lon2)

    try:
        start_node = ox.distance.nearest_nodes(G, lon1, lat1)
        end_node = ox.distance.nearest_nodes(G, lon2, lat2)

        distance_meters = nx.astar_path_length(G, start_node, end_node, weight="travel_cost")
        return round(distance_meters / 1000, 2)

    except Exception as e:
        print("A* failed, fallback:", e)
        return haversine(lat1, lon1, lat2, lon2)
    
    from itertools import islice

    def get_alternative_routes(lat1, lon1, lat2, lon2):

        start_node = ox.distance.nearest_nodes(G, lon1, lat1)
        end_node = ox.distance.nearest_nodes(G, lon2, lat2)

        routes = list(
            islice(
                nx.shortest_simple_paths(
                    G,
                    start_node,
                    end_node,
                    weight="travel_cost"
                ),
                3
            )
        )

        return routes

    print("Running A* routing...")

    start_node = ox.distance.nearest_nodes(G, lon1, lat1)
    end_node = ox.distance.nearest_nodes(G, lon2, lat2)

    distance_meters = nx.astar_path_length(G, start_node, end_node, weight="length")

    distance_km = distance_meters / 1000

    print(f"A* distance: {distance_km:.2f} km")

    return distance_km

# =========================================================
# HOME + LOGIN
# =========================================================

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        login_identifier = request.form.get("login_identifier", "").strip()
        password = request.form.get("password", "")

        if not login_identifier or not password:
            return render_template(
                "login.html",
                login_error="Please enter your username/email and password."
            )

        users = load_users()
        matched_user = None

        for user in users:
            username = str(user.get("username") or "").strip().lower()
            email = str(user.get("email") or "").strip().lower()

            if login_identifier.lower() in [username, email]:
                matched_user = user
                break

        if matched_user is None:
            return render_template(
                "login.html",
                login_error="Invalid username/email or password."
            )

        if str(matched_user.get("account_status") or "").strip().lower() != "active":
            return render_template(
                "login.html",
                login_error="Your account is not active. Please contact the administrator."
            )

        password_hash = str(matched_user.get("password_hash") or "")

        if not password_hash or not check_password_hash(password_hash, password):
            return render_template(
                "login.html",
                login_error="Invalid username/email or password."
            )

        # Each browser receives its own independent Flask session.
        session.clear()

        session["user_id"] = str(matched_user.get("user_id") or "")
        session["first_name"] = str(matched_user.get("first_name") or "")
        session["last_name"] = str(matched_user.get("last_name") or "")
        session["username"] = str(matched_user.get("username") or "")

        session["user_name"] = (
            f"{matched_user.get('first_name', '')} "
            f"{matched_user.get('last_name', '')}"
        ).strip()

        session["user_phone"] = str(matched_user.get("mobile") or "")
        session["user_email"] = str(matched_user.get("email") or "")
        session["role"] = str(matched_user.get("role") or "").strip().lower()
        session["stp_id"] = safe_user_value(matched_user, "stp_id")
        session["tanker_operator_id"] = safe_user_value(matched_user, "tanker_operator_id")

        # Keep the existing buyer session variables.
        if session["role"] == "demand":
            session["buyer_name"] = session["user_name"]
            session["buyer_phone"] = session["user_phone"]
            return redirect(url_for("demand"))

        if session["role"] == "stp":

            stp_id = str(session.get("stp_id") or "").strip()

            if not stp_id:
                session.clear()

                return render_template(
                    "login.html",
                    login_error="No STP is assigned to this account."
                )

            return redirect(
                url_for(
                    "supply",
                    stp_id=stp_id
                )
            )

        if session["role"] == "tanker":
            # Keep the existing tanker dashboard flow, but use the
            # registered Tanker Operator ID linked during signup.
            session["tanker_operator_id"] = str(
                matched_user.get("tanker_operator_id") or ""
            ).strip()
            session["tanker_operator_name"] = session["user_name"]
            return redirect(url_for("tanker_dashboard"))

        if session["role"] == "admin":
            return redirect(url_for("admin_dashboard"))

        session.clear()

        return render_template(
            "login.html",
            login_error="Your account has an invalid role."
        )

    return render_template("login.html")


@app.route('/signup', methods=['GET', 'POST'])
def signup():

    if request.method == 'POST':

        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        mobile = request.form.get("mobile", "").strip()
        email = request.form.get("email", "").strip().lower()
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        role = request.form.get("role", "").strip().lower()

        # Role-specific identity fields from signup.html.
        stp_id = request.form.get("stp_id", "").strip()
        tanker_operator_id = request.form.get("tanker_id", "").strip()

        allowed_roles = {"demand", "stp", "tanker", "admin"}

        if not all([
            first_name,
            last_name,
            mobile,
            email,
            username,
            password,
            confirm_password,
            role
        ]):
            return render_template(
                "signup.html",
                signup_error="Please fill in all fields."
            )

        if role not in allowed_roles:
            return render_template(
                "signup.html",
                signup_error="Please select a valid account type."
            )

        # STP operators must provide an existing STP ID.
        if role == "stp":
            if not stp_id:
                return render_template(
                    "signup.html",
                    signup_error="Please enter your STP ID."
                )

            stp_exists = any(
                str(stp.get("stp_id") or "").strip().lower() == stp_id.lower()
                for stp in load_stps()
            )

            if not stp_exists:
                return render_template(
                    "signup.html",
                    signup_error="Invalid STP ID. Please enter a registered STP ID."
                )

        # Tanker operators must provide an existing tanker operator ID.
        if role == "tanker":
            if not tanker_operator_id:
                return render_template(
                    "signup.html",
                    signup_error="Please enter your Tanker Operator ID."
                )

            tanker_exists = False
            if os.path.exists(TANKER_REGISTRATIONS_FILE):
                try:
                    with open(
                        TANKER_REGISTRATIONS_FILE,
                        "r",
                        newline="",
                        encoding="utf-8"
                    ) as f:
                        reader = csv.DictReader(f)
                        tanker_exists = any(
                            str(row.get("operator_id") or "").strip().lower() == tanker_operator_id.lower()
                            for row in reader
                        )
                except Exception as e:
                    print("Tanker operator ID validation failed:", e)

            if not tanker_exists:
                return render_template(
                    "signup.html",
                    signup_error="Invalid Tanker Operator ID. Please enter a registered operator ID."
                )

        if password != confirm_password:
            return render_template(
                "signup.html",
                signup_error="Passwords do not match."
            )

        if len(password) < 8:
            return render_template(
                "signup.html",
                signup_error="Password must be at least 8 characters long."
            )

        users = load_users()

        for user in users:
            existing_username = str(user.get("username") or "").strip().lower()
            existing_email = str(user.get("email") or "").strip().lower()

            if username == existing_username:
                return render_template(
                    "signup.html",
                    signup_error="That username is already registered."
                )

            if email == existing_email:
                return render_template(
                    "signup.html",
                    signup_error="That email address is already registered."
                )

            if mobile == str(user.get("mobile") or "").strip():
                return render_template(
                    "signup.html",
                    signup_error="That mobile number is already registered."
                )

        user_id = "USR-" + uuid.uuid4().hex[:10].upper()

        new_user = {
            "user_id": user_id,
            "first_name": first_name,
            "last_name": last_name,
            "username": username,
            "mobile": mobile,
            "email": email,
            "password_hash": generate_password_hash(password),
            "role": role,
            "stp_id": stp_id if role == "stp" else "",
            "tanker_operator_id": tanker_operator_id if role == "tanker" else "",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "account_status": "active"
        }

        append_user(new_user)

        return redirect(
            url_for(
                "login",
                signup_success="Account created successfully. Please log in."
            )
        )

    stps = load_stps()
    return render_template(
        "signup.html",
        stps=stps
    )


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/delete_account", methods=["POST"])
def delete_account():

    # User must be logged in
    if not session.get("user_id"):
        return redirect(url_for("login"))

    user_id = str(session.get("user_id") or "")

    if not user_id:
        return redirect(url_for("login"))

    # Make sure users.xlsx exists
    ensure_users_file()

    with users_lock:

        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

        # Find user_id column
        if "user_id" not in headers:
            workbook.close()
            return "User ID column not found in users.xlsx", 500

        user_id_column = headers.index("user_id") + 1

        user_found = False

        # Find and delete the logged-in user's row
        for row in range(2, sheet.max_row + 1):

            current_user_id = str(
                sheet.cell(
                    row=row,
                    column=user_id_column
                ).value or ""
            ).strip()

            if current_user_id == user_id:

                sheet.delete_rows(row, 1)
                user_found = True
                break

        workbook.save(USERS_FILE)
        workbook.close()

    # Clear the current login session
    session.clear()

    if user_found:
        return redirect(
            url_for(
                "login",
                account_deleted="Account deleted successfully."
            )
        )

    return redirect(url_for("login"))


# =========================================================
# CURRENT LOGGED-IN USER
# =========================================================

@app.route("/profile")
def profile():

    # Check whether a user is logged in
    if not session.get("user_id"):
        return redirect(url_for("login"))

    # Get the complete user record from users.xlsx
    users = load_users()
    logged_in_user = None

    for user in users:
        if str(user.get("user_id") or "") == str(session.get("user_id") or ""):
            logged_in_user = user
            break

    if logged_in_user is None:
        session.clear()
        return redirect(url_for("login"))

    return render_template(
        "profile.html",
        user={
            "user_id": logged_in_user.get("user_id", ""),
            "first_name": logged_in_user.get("first_name", ""),
            "last_name": logged_in_user.get("last_name", ""),
            "name": f"{logged_in_user.get('first_name', '')} {logged_in_user.get('last_name', '')}".strip(),
            "username": logged_in_user.get("username", ""),
            "mobile": logged_in_user.get("mobile", ""),
            "email": logged_in_user.get("email", ""),
            "role": logged_in_user.get("role", ""),
            "created_at": logged_in_user.get("created_at", ""),
            "account_status": logged_in_user.get("account_status", "")
        }
    )
    
@app.route("/api/current_user")
def current_user():
    """Return only the user stored in this browser's Flask session."""
    user_id = session.get("user_id")

    if not user_id:
        return jsonify({
            "logged_in": False,
            "initials": "👤",
            "name": "Guest",
            "role": ""
        })

    first_name = str(session.get("first_name") or "").strip()
    last_name = str(session.get("last_name") or "").strip()

    initials = ""
    if first_name:
        initials += first_name[0].upper()
    if last_name:
        initials += last_name[0].upper()
    if not initials:
        initials = "👤"

    return jsonify({
        "logged_in": True,
        "first_name": first_name,
        "last_name": last_name,
        "name": str(session.get("user_name") or "").strip(),
        "role": str(session.get("role") or "").strip(),
        "initials": initials
    })


@app.route("/tanker/register")
def tanker_register():
    return render_template("tanker_register.html")

# =========================================================
# STP REGISTRATION
# =========================================================

@app.route("/stp/register", methods=["GET", "POST"])
def stp_register():

    if request.method == "GET":
        return render_template("stp_register.html")

    # -----------------------------
    # Read submitted form data
    # -----------------------------

    owner_name = request.form.get("owner_name", "").strip()
    phone = request.form.get("phone", "").strip()
    email = request.form.get("email", "").strip()
    company_name = request.form.get("company_name", "").strip()

    stp_name = request.form.get("stp_name", "").strip()
    technology = request.form.get("technology", "").strip()

    total_capacity_kld = request.form.get(
        "total_capacity_kld", "0"
    )

    current_load_kld = request.form.get(
        "current_load_kld", "0"
    )

    treatment_cost_per_kl = request.form.get(
        "treatment_cost_per_kl", "0"
    )

    quality_grade = request.form.get(
        "quality_grade", ""
    ).strip()

    latitude = request.form.get(
        "latitude", ""
    ).strip()

    longitude = request.form.get(
        "longitude", ""
    ).strip()


    # -----------------------------
    # Basic validation
    # -----------------------------

    if not owner_name:
        return "Owner name is required", 400

    if not phone:
        return "Phone number is required", 400

    if not email:
        return "Email is required", 400

    if not stp_name:
        return "STP name is required", 400

    if not technology:
        return "STP technology is required", 400

    if not latitude or not longitude:
        return "STP location is required", 400


    # -----------------------------
    # Convert numerical values
    # KLD → MLD
    # -----------------------------

    try:

        total_capacity_mld = (
            float(total_capacity_kld) / 1000
        )

        current_load_mld = (
            float(current_load_kld) / 1000
        )

        treatment_cost = float(
            treatment_cost_per_kl
        )

    except ValueError:

        return "Invalid numerical value submitted", 400


    # -----------------------------
    # Validate capacity
    # -----------------------------

    if total_capacity_mld <= 0:
        return "Total capacity must be greater than zero", 400

    if current_load_mld < 0:
        return "Current load cannot be negative", 400

    if current_load_mld > total_capacity_mld:
        return (
            "Current load cannot exceed total capacity",
            400
        )


    # -----------------------------
    # Generate registration ID
    # -----------------------------

    registration_id = (
        "REG-" +
        datetime.now().strftime("%Y%m%d%H%M%S")
    )


    # -----------------------------
    # Registration record
    # -----------------------------

    registration = {
        "registration_id": registration_id,
        "stp_id": "",
        "owner_name": owner_name,
        "phone": phone,
        "email": email,
        "company_name": company_name,
        "stp_name": stp_name,
        "latitude": latitude,
        "longitude": longitude,
        "technology": technology,
        "total_capacity_mld": total_capacity_mld,
        "current_load_mld": current_load_mld,
        "treatment_cost_per_kl": treatment_cost,
        "quality_grade": quality_grade,
        "verification_status": "pending",
        "registration_date": datetime.now().isoformat(),
        "approved_at": ""
    }


    # -----------------------------
    # Save registration
    # -----------------------------

    file_exists = os.path.exists(
        STP_REGISTRATIONS_FILE
    )

    with open(
        STP_REGISTRATIONS_FILE,
        "a",
        newline="",
        encoding="utf-8"
    ) as f:

        fieldnames = [
            "registration_id",
            "stp_id",
            "owner_name",
            "phone",
            "email",
            "company_name",
            "stp_name",
            "latitude",
            "longitude",
            "technology",
            "total_capacity_mld",
            "current_load_mld",
            "treatment_cost_per_kl",
            "quality_grade",
            "verification_status",
            "registration_date",
            "approved_at"
        ]

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames
        )

        if not file_exists:
            writer.writeheader()

        writer.writerow(registration)


    return render_template(
        "stp_registration_success.html",
        registration_id=registration_id,
        stp_name=stp_name
    )


@app.route("/tanker/status", methods=["GET", "POST"])
def tanker_status():

    if request.method == "POST":

        operator_id = request.form.get("operator_id", "").strip()
        phone = request.form.get("phone", "").strip()

        operator = None

        if os.path.exists(TANKER_REGISTRATIONS_FILE):

            with open(
                TANKER_REGISTRATIONS_FILE,
                "r",
                newline="",
                encoding="utf-8"
            ) as f:

                reader = csv.DictReader(f)

                for row in reader:

                    if (
                        row.get("operator_id", "").strip() == operator_id
                        and
                        row.get("phone", "").strip() == phone
                    ):
                        operator = row
                        break

        return render_template(
            "tanker_status.html",
            operator=operator,
            searched=True
        )

    return render_template(
        "tanker_status.html",
        operator=None,
        searched=False
    )

@app.route("/tanker/register/contracted", methods=["GET", "POST"])
def tanker_register_contracted():

    if request.method == "POST":

        # Operator details
        operator_name = request.form.get("operator_name")
        phone = request.form.get("phone")
        email = request.form.get("email")

        # Contract details
        contract_id = request.form.get("contract_id")
        contract_start = request.form.get("contract_start")
        contract_end = request.form.get("contract_end")

        # Tanker details
        registration_no = request.form.get("registration_no")
        capacity = request.form.get("capacity")
        vehicle_model = request.form.get("vehicle_model")

        # CSV file location
        csv_file = os.path.join(
            app.root_path,
            "database",
            "tanker_registrations.csv"
        )

        # Generate operator ID
        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
            existing_df = pd.read_csv(csv_file)
            operator_number = len(existing_df) + 1
        else:
            operator_number = 1

        operator_id = f"OP-BLR-{operator_number:04d}"

        # Create registration record
        new_operator = {
            "operator_id": operator_id,
            "operator_name": operator_name,
            "operator_type": "contracted",
            "phone": phone,
            "email": email,
            "area": "",
            "pincode": "",
            "contract_id": contract_id,
            "contract_start": contract_start,
            "contract_end": contract_end,
            "tanker_registration_no": registration_no,
            "tanker_capacity_kl": capacity,
            "vehicle_model": vehicle_model,
            "water_type_supported": "",
            "service_radius_km": "",
            "verification_status": "pending",
            "registration_date": date.today().isoformat()
        }

        # Convert to DataFrame
        new_df = pd.DataFrame([new_operator])

        # Save to CSV
        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
            new_df.to_csv(
                csv_file,
                mode="a",
                header=False,
                index=False
            )
        else:
            new_df.to_csv(
                csv_file,
                mode="w",
                header=True,
                index=False
            )

        print("NEW CONTRACTED TANKER OPERATOR REGISTERED")
        print(new_operator)
        print("DATA SAVED TO:", csv_file)

        return render_template(
            "registration_success.html",
        operator_id=operator_id,
        operator_type="Existing Purvankara Partner"
)

    return render_template("tanker_register_contracted.html")

@app.route("/tanker/register/independent", methods=["GET", "POST"])
def tanker_register_independent():

    if request.method == "POST":

        # Get the submitted form data
        operator_name = request.form.get("operator_name")
        phone = request.form.get("phone")
        email = request.form.get("email")

        area = request.form.get("area")
        pincode = request.form.get("pincode")

        registration_no = request.form.get("registration_no")
        capacity = request.form.get("capacity")
        vehicle_model = request.form.get("vehicle_model")

        water_type = request.form.get("water_type")
        radius = request.form.get("radius")

        # CSV file location
        csv_file = os.path.join(
            app.root_path,
            "database",
            "tanker_registrations.csv"
        )
        print("CSV PATH:", csv_file)
        print("CSV EXISTS:", os.path.exists(csv_file))

        # Generate a new operator ID
        if os.path.exists(csv_file):

            existing_df = pd.read_csv(csv_file)

            operator_number = len(existing_df) + 1

        else:
            operator_number = 1

        operator_id = f"OP-BLR-{operator_number:04d}"

        # Create the new registration
        new_operator = {
            "operator_id": operator_id,
            "operator_name": operator_name,
            "operator_type": "independent",
            "phone": phone,
            "email": email,
            "area": area,
            "pincode": pincode,
            "contract_id": "",
            "contract_start": "",
            "contract_end": "",
            "tanker_registration_no": registration_no,
            "tanker_capacity_kl": capacity,
            "vehicle_model": vehicle_model,
            "water_type_supported": water_type,
            "service_radius_km": radius,
            "verification_status": "pending",
            "registration_date": date.today().isoformat()
        }

        # Add the registration to the CSV
        new_df = pd.DataFrame([new_operator])

        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
                new_df.to_csv(
                    csv_file,
                    mode="a",
                    header=False,
                    index=False
                )
        else:
                new_df.to_csv(
                    csv_file,
                    mode="w",
                    header=True,
                    index=False
                )

        print("DATA SAVED TO:", csv_file)

        print("NEW TANKER OPERATOR REGISTERED")
        print(new_operator)

        return render_template(
            "registration_success.html",
        operator_id=operator_id,
        operator_type="Independent Operator"
)

    return render_template("tanker_register_independent.html")


@app.route("/admin")
def admin_dashboard():

    # =========================
    # LOAD STPs
    # =========================

    stps = load_stps()


    # =========================
    # LOAD ORDERS
    # =========================

    orders = []

    if os.path.exists(ORDERS_FILE):

        with open(
            ORDERS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)
            orders = list(reader)


    # =========================
    # LOAD TANKER REGISTRATIONS
    # =========================

    tanker_operators = []

    if os.path.exists(TANKER_REGISTRATIONS_FILE):

        with open(
            TANKER_REGISTRATIONS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)
            tanker_operators = list(reader)


    total_tanker_operators = len(
        tanker_operators
    )


    pending_tanker_operators = sum(
        1
        for operator in tanker_operators
        if operator.get(
            "verification_status",
            ""
        ).strip().lower() == "pending"
    )


    approved_tanker_operators = sum(
        1
        for operator in tanker_operators
        if operator.get(
            "verification_status",
            ""
        ).strip().lower() == "approved"
    )


    rejected_tanker_operators = sum(
        1
        for operator in tanker_operators
        if operator.get(
            "verification_status",
            ""
        ).strip().lower() == "rejected"
    )


    # =========================
    # LOAD STP REGISTRATIONS
    # =========================

    stp_registrations = []

    if os.path.exists(STP_REGISTRATIONS_FILE):

        with open(
            STP_REGISTRATIONS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)
            stp_registrations = list(reader)


    # =========================
    # ADMIN PAGE
    # =========================

    return render_template(
        "admin.html",

        stps=stps,

        orders=orders,

        tanker_operators=tanker_operators,

        total_tanker_operators=
            total_tanker_operators,

        pending_tanker_operators=
            pending_tanker_operators,

        approved_tanker_operators=
            approved_tanker_operators,

        rejected_tanker_operators=
            rejected_tanker_operators,

        stp_registrations=
            stp_registrations
    )


@app.route("/admin/tanker/<operator_id>/status/<status>")
def update_tanker_status(operator_id, status):

    # Only allow valid statuses
    if status not in ["approved", "rejected"]:
        return redirect("/admin")

    # Make sure the tanker registration file exists
    if not os.path.exists(TANKER_REGISTRATIONS_FILE):
        return redirect("/admin")

    rows = []

    # Read all existing tanker registrations
    with open(
        TANKER_REGISTRATIONS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    # If the CSV is empty or damaged
    if not fieldnames:
        return redirect("/admin")

    operator_found = False

    # Update the selected tanker operator
    for operator in rows:

        if operator.get("operator_id") == operator_id:

            operator["verification_status"] = status

            operator_found = True

            break

    # If operator ID does not exist
    if not operator_found:
        return redirect("/admin")

    # Save the entire CSV again
    with open(
        TANKER_REGISTRATIONS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames
        )

        writer.writeheader()
        writer.writerows(rows)

    return redirect("/admin")

@app.route("/api/stp_orders")
def api_stp_orders():
    # Return only orders assigned to the STP operator's selected STP.
    if session.get("role") != "stp":
        return jsonify({"error": "Unauthorized"}), 403

    requested_stp_id = (request.args.get("stp_id") or "").strip()

    results = []

    if not os.path.exists(ORDERS_FILE):
        return jsonify(results)

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for row in reader:
            row_stp_id = (row.get("stp_id") or "").strip()

            if requested_stp_id and row_stp_id != requested_stp_id:
                continue

            results.append({
                "order_id": row.get("order_id", ""),
                "stp_id": row.get("stp_id", ""),
                "stp_name": row.get("stp_name", ""),
                "quantity_kld": row.get("quantity_kld", ""),
                "quality": row.get("quality", ""),
                "water_type": row.get("water_type", ""),
                "distance_km": row.get("distance_km", ""),
                "location": row.get("location", ""),
                "buyer_name": row.get("buyer_name", ""),
                "buyer_phone": row.get("buyer_phone", ""),
                "status": row.get("status", ""),
                "created_at": row.get("created_at", ""),
                "payment_status": row.get("payment_status", ""),
                "accepted_at": row.get("accepted_at", ""),
                "stp_latitude": row.get("stp_latitude", ""),
                "stp_longitude": row.get("stp_longitude", ""),
                "delivery_latitude": row.get("delivery_latitude", ""),
                "delivery_longitude": row.get("delivery_longitude", "")
            })

    results.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return jsonify(results)


@app.route("/api/stp_order_tracking/<order_id>")
def stp_order_tracking(order_id):
    # Return one order for the STP operator tracking page.
    if session.get("role") != "stp":
        return jsonify({"error": "Unauthorized"}), 403

    requested_stp_id = (request.args.get("stp_id") or "").strip()

    if not os.path.exists(ORDERS_FILE):
        return jsonify({"error": "Orders file not found"}), 404

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for row in reader:
            if (row.get("order_id") or "").strip() != order_id.strip():
                continue

            row_stp_id = (row.get("stp_id") or "").strip()

            if requested_stp_id and row_stp_id != requested_stp_id:
                return jsonify({"error": "Order does not belong to this STP"}), 403

            return jsonify(row)

    return jsonify({"error": "Order not found"}), 404


@app.route("/api/stps")
def api_stps():
    return jsonify(load_stps())

@app.route("/admin/stp/<registration_id>/status/<status>")
def update_stp_status(registration_id, status):

    # =========================
    # LOAD REGISTRATIONS
    # =========================

    if not os.path.exists(STP_REGISTRATIONS_FILE):
        return redirect("/admin")

    rows = []

    with open(
        STP_REGISTRATIONS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)

        fieldnames = reader.fieldnames or []

        for row in reader:
            rows.append(row)


    # =========================
    # FIND REGISTRATION
    # =========================

    registration = None

    for row in rows:

        if row.get("registration_id", "") == registration_id:

            registration = row
            break


    if registration is None:
        return redirect("/admin")


    # =========================
    # REJECT
    # =========================

    if status == "rejected":

        registration["verification_status"] = "rejected"

        with open(
            STP_REGISTRATIONS_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as f:

            writer = csv.DictWriter(
                f,
                fieldnames=fieldnames
            )

            writer.writeheader()
            writer.writerows(rows)

        return redirect("/admin")


    # =========================
    # APPROVE
    # =========================

    if status != "approved":
        return "Invalid status", 400


    stps = load_stps()


    # =========================
    # GENERATE NEXT STP ID
    # =========================

    highest_id = 0

    for stp in stps:

        stp_id = str(
            stp.get("stp_id", "")
        ).strip()

        if stp_id.startswith("PSTP"):

            try:

                number = int(
                    stp_id.replace("PSTP", "")
                )

                highest_id = max(
                    highest_id,
                    number
                )

            except ValueError:
                pass


    new_stp_id = f"PSTP{highest_id + 1:03d}"


    # =========================
    # CONVERT VALUES
    # =========================

    try:

        total_capacity = float(
            registration.get(
                "total_capacity_mld",
                0
            )
        )

        current_load = float(
            registration.get(
                "current_load_mld",
                0
            )
        )

        treatment_cost = float(
            registration.get(
                "treatment_cost_per_kl",
                0
            )
        )

        latitude = float(
            registration.get(
                "latitude",
                0
            )
        )

        longitude = float(
            registration.get(
                "longitude",
                0
            )
        )

    except (ValueError, TypeError):

        return "Invalid STP registration data", 400


    # =========================
    # AVAILABLE CAPACITY
    # =========================

    available_capacity = (
        total_capacity - current_load
    )


    # =========================
    # CREATE STP
    # =========================

    now = datetime.now()

    new_stp = {

        "stp_id": new_stp_id,

        "stp_name": registration.get(
            "stp_name",
            ""
        ),

        "latitude": latitude,

        "longitude": longitude,

        "technology": registration.get(
            "technology",
            ""
        ),

        "total_capacity_mld": total_capacity,

        "current_load_mld": current_load,

        "available_capacity_mld":
            available_capacity,

        "treatment_cost_per_kl":
            treatment_cost,

        "quality_grade": registration.get(
            "quality_grade",
            "General"
        ),

        "last_reset_date":
            now.strftime("%Y-%m-%d"),

        "last_reset_at":
            now.isoformat()

    }


    # =========================
    # ADD STP TO JSON
    # =========================

    stps.append(new_stp)

    save_stps(stps)


    # =========================
    # UPDATE REGISTRATION
    # =========================

    registration["stp_id"] = new_stp_id

    registration["verification_status"] = "approved"

    registration["approved_at"] = now.isoformat()


    # =========================
    # SAVE REGISTRATION CSV
    # =========================

    with open(
        STP_REGISTRATIONS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames
        )

        writer.writeheader()
        writer.writerows(rows)


    return redirect("/admin")

# =========================
# ADD STP
# =========================
@app.route("/admin/add_stp", methods=["POST"])
def add_stp():
    stps = load_stps()

    new_stp = {
        "stp_id": request.form["id"],
        "stp_name": request.form["name"],
        "latitude": float(request.form["lat"]),
        "longitude": float(request.form["lon"]),
        "technology": "Manual",
        "total_capacity_mld": float(request.form["capacity"]),
        "current_load_mld": 0.0,
        "available_capacity_mld": float(request.form["capacity"]),
        "treatment_cost_per_kl": 5.0,
        "quality_grade": "General",

        "last_reset_date": date.today().isoformat(),
        "last_reset_at": datetime.now().isoformat()
    }

    stps.append(new_stp)
    save_stps(stps)

    return redirect(url_for("admin_dashboard"))


# =========================
# DELETE STP
# =========================
@app.route("/admin/delete_stp/<stp_id>")
def delete_stp(stp_id):
    stps = load_stps()

    stps = [s for s in stps if str(s["stp_id"]) != str(stp_id)]

    save_stps(stps)

    return redirect(url_for("admin_dashboard"))

# =========================================================
# DEMAND SIDE
# =========================================================

@app.route('/demand')
def demand():
    payment_success = request.args.get("payment_success")

    return render_template(
        "demand.html",
        payment_success=payment_success
    )

@app.route("/track")
def track_page():
    return render_template("track.html")



# =========================================================
# DEMAND HEATMAP API
# =========================================================

@app.route("/api/demand_heatmap")
def demand_heatmap():

    demand_data = []

    if not os.path.exists(DEMAND_CSV_FILE):
        return jsonify({
            "error": "synthetic_orders.csv not found"
        }), 404

    try:

        with open(
            DEMAND_CSV_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as file:

            reader = csv.DictReader(file)

            for row in reader:

                try:

                    latitude = float(row["latitude"])
                    longitude = float(row["longitude"])
                    quantity = float(row["quantity_kld"])

                    demand_data.append({
                        "latitude": latitude,
                        "longitude": longitude,
                        "quantity_kld": quantity
                    })

                except (
                    KeyError,
                    ValueError,
                    TypeError
                ):
                    # Ignore malformed rows
                    continue

        return jsonify(demand_data)

    except Exception as e:

        print("Demand heatmap error:", e)

        return jsonify({
            "error": "Could not read synthetic_orders.csv"
        }), 500
    

@app.route("/api/search_place")
def api_search_place():

    auto_reset_capacity()

    place = request.args.get("place")
    lat = request.args.get("lat")
    lon = request.args.get("lon")

    # Keep the existing location behavior: typed location or live location.
    if lat and lon:
        lat = float(lat)
        lon = float(lon)

        reverse_url = (
            f"https://nominatim.openstreetmap.org/reverse"
            f"?format=json&lat={lat}&lon={lon}"
        )
        try:
            response = requests.get(
                reverse_url,
                headers={"User-Agent": "wastewater-app"},
                timeout=5
            )
            data = response.json()
        except Exception as e:
            print("Reverse API failed:", e)
            data = {}
        
        address = data.get("address", {})

        location_name = format_clean_address(address, lat, lon)

        # fallback (if still empty)
        if not location_name or location_name.strip() == "":
            location_name = data.get("display_name", f"{lat}, {lon}")

        print("Using LIVE coordinates:", lat, lon)

    elif place and place != "Using Live Location":
        geo_url = f"https://nominatim.openstreetmap.org/search?format=json&q={place}, Bangalore"
        response = requests.get(geo_url, headers={"User-Agent":"wastewater-app"})
        geo_data = response.json()

        if not geo_data:
            return jsonify({"error":"Place not found"}), 404

        lat = float(geo_data[0]["lat"])
        lon = float(geo_data[0]["lon"])
        location_name = place

    else:
        return jsonify({"error": "No location provided"}), 400

    # 🔥 ADD THIS BLOCK HERE (VERY IMPORTANT)

    required_kld_raw = request.args.get("required_kld")
    required_kld = float(required_kld_raw) if required_kld_raw and required_kld_raw.strip() != "" else 0

    required_quality = request.args.get("quality")
    required_type = request.args.get("type")

    stps = load_stps()
    nearby = []

    for stp in stps:

        if not stp.get("latitude") or not stp.get("longitude"):
            continue

        # FILTER BY QUALITY
        if required_quality and stp.get("quality_grade") != required_quality:
            continue
        
        # FILTER BY TYPE (SAFE FIX)
        if required_type and stp.get("water_type") and stp.get("water_type") != required_type:
            continue

        # Stage 1: Fast filtering
        approx_distance = haversine(lat, lon, stp["latitude"], stp["longitude"])

        if approx_distance > 100:
            continue

        # Stage 2: Accurate routing
        distance = astar_distance(lat, lon, stp["latitude"], stp["longitude"])

        if distance > 100:
            continue

        stp_copy = stp.copy()
        stp_copy["distance_km"] = round(distance,2)
        nearby.append(stp_copy)

    nearby.sort(key=lambda x: x["distance_km"])
    nearest = nearby[0] if nearby else None
    
    if not nearest:
        return jsonify({
        "searched_location": {
            "name": location_name,
            "latitude": lat,
            "longitude": lon
        },
        "nearest_stp": None,
        "all_stps": []
    })

    return jsonify({
        "searched_location": {
            "name": location_name,
            "latitude": lat,
            "longitude": lon
        },
        "nearest_stp": nearest,
        "all_stps": [s for s in stps if s.get("latitude") and s.get("longitude")]
    })

@app.route("/create_order", methods=["POST"])
def create_order():
    data = request.json or {}

    required = ["stp_id", "stp_name", "quantity_kld", "quality", "water_type", "distance_km", "location"]
    missing = [key for key in required if key not in data]
    if missing:
        return jsonify({"error": "Missing fields", "fields": missing}), 400

    order_id = "ORD-" + uuid.uuid4().hex[:10].upper()

    row = {
        "order_id": order_id,
        "stp_id": data["stp_id"],
        "stp_name": data["stp_name"],
        "quantity_kld": data["quantity_kld"],
        "quality": data["quality"],
        "water_type": data["water_type"],
        "distance_km": data["distance_km"],
        "location": data["location"],
        "buyer_user_id": session.get("user_id") or "",
        "buyer_name": session.get("buyer_name") or session.get("user_name") or "Unknown",
        "buyer_phone": session.get("buyer_phone") or session.get("user_phone") or "N/A",
        "status": "Pending",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "payment_status": "Pending",
        "accepted_at": "",
        "capacity_release_at": "",
        "capacity_released": "False"
    }

    with open(ORDERS_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writerow(row)

    return jsonify({"message": "Order created successfully", "order_id": order_id})

# =========================================================
# REORDER EXISTING ORDER
# =========================================================

@app.route("/reorder/<order_id>", methods=["POST"])
def reorder_order(order_id):

    # -----------------------------------------------------
    # USER MUST BE LOGGED IN
    # -----------------------------------------------------

    user_id = session.get("user_id")

    buyer_name = (
        session.get("buyer_name")
        or session.get("user_name")
    )

    buyer_phone = (
        session.get("buyer_phone")
        or session.get("user_phone")
    )

    if not user_id:
        return jsonify({
            "success": False,
            "error": "Please log in to reorder."
        }), 401


    # -----------------------------------------------------
    # ONLY DEMAND USERS CAN REORDER
    # -----------------------------------------------------

    if str(session.get("role") or "").lower() != "demand":
        return jsonify({
            "success": False,
            "error": "Only demand users can reorder."
        }), 403


    # -----------------------------------------------------
    # CHECK ORDERS FILE
    # -----------------------------------------------------

    if not os.path.exists(ORDERS_FILE):
        return jsonify({
            "success": False,
            "error": "Orders file not found."
        }), 404


    # -----------------------------------------------------
    # FIND ORIGINAL ORDER
    # -----------------------------------------------------

    original_order = None

    with open(
        ORDERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            if (
                str(row.get("order_id", "")).strip()
                ==
                str(order_id).strip()
            ):

                # -----------------------------------------
                # VERIFY THAT THIS ORDER BELONGS
                # TO THE CURRENT LOGGED-IN USER
                # -----------------------------------------

                matches_user = (
                    user_id
                    and
                    row.get("buyer_user_id", "") == user_id
                )

                matches_legacy = (
                    not row.get("buyer_user_id", "")
                    and buyer_name
                    and buyer_phone
                    and row.get("buyer_name") == buyer_name
                    and row.get("buyer_phone") == buyer_phone
                )

                if not (
                    matches_user
                    or matches_legacy
                ):

                    return jsonify({
                        "success": False,
                        "error": "You cannot reorder another user's order."
                    }), 403

                original_order = row
                break


    # -----------------------------------------------------
    # ORDER NOT FOUND
    # -----------------------------------------------------

    if original_order is None:

        return jsonify({
            "success": False,
            "error": "Original order not found."
        }), 404


    # -----------------------------------------------------
    # GENERATE NEW ORDER ID
    # -----------------------------------------------------

    new_order_id = (
        "ORD-" +
        uuid.uuid4().hex[:10].upper()
    )


    # -----------------------------------------------------
    # CREATE NEW ORDER USING OLD ORDER DETAILS
    # -----------------------------------------------------

    new_order = {

        "order_id":
            new_order_id,

        "stp_id":
            original_order.get("stp_id", ""),

        "stp_name":
            original_order.get("stp_name", ""),

        "quantity_kld":
            original_order.get("quantity_kld", ""),

        "quality":
            original_order.get("quality", ""),

        "water_type":
            original_order.get("water_type", ""),

        "distance_km":
            original_order.get("distance_km", ""),

        "location":
            original_order.get("location", ""),

        # Always use CURRENT logged-in account
        "buyer_user_id":
            user_id,

        "buyer_name":
            buyer_name or "Unknown",

        "buyer_phone":
            buyer_phone or "N/A",

        # Reset order state
        "status":
            "Pending",

        "created_at":
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

        # Payment must be selected again
        "payment_status":
            "Pending",

        # Old fulfilment data must NOT be copied
        "accepted_at":
            "",

        "capacity_release_at":
            "",

        "capacity_released":
            "False"
    }


    # -----------------------------------------------------
    # SAVE NEW ORDER
    # -----------------------------------------------------

    with open(
        ORDERS_FILE,
        "a",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=ORDER_FIELDS
        )

        writer.writerow(new_order)


    # -----------------------------------------------------
    # RETURN NEW ORDER ID
    # -----------------------------------------------------

    return jsonify({

        "success": True,

        "message":
            "Order recreated successfully.",

        "original_order_id":
            order_id,

        "new_order_id":
            new_order_id
    })


@app.route("/invoice")
def invoice():

    order_id = request.args.get("order_id")

    if not order_id:
        return "Order ID is missing", 400

    order = None

    if os.path.exists(ORDERS_FILE):

        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:

            reader = csv.DictReader(f)

            for row in reader:

                if row.get("order_id") == order_id:
                    order = row
                    break

    if not order:
        return "Order not found", 404

    # Only allow the logged-in buyer to view their own invoice.
    current_user_id = session.get("user_id")
    current_buyer_name = session.get("buyer_name") or session.get("user_name")
    current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

    authorized = (
        current_user_id and
        order.get("buyer_user_id", "") == current_user_id
    ) or (
        not order.get("buyer_user_id", "") and
        current_buyer_name and
        current_buyer_phone and
        order.get("buyer_name") == current_buyer_name and
        order.get("buyer_phone") == current_buyer_phone
    )

    if not authorized:
        return "Unauthorized", 403

    # Convert the existing order data
    # into the names expected by invoice.html

    # =========================================================
    # INVOICE CALCULATION
    # =========================================================

    quantity = float(order.get("quantity_kld") or 0)

    # Price of treated wastewater per KL
    WATER_RATE = 30.0

    # Transportation charge per KL
    TRANSPORT_RATE = 10.0

    # GST rate
    GST_RATE = 0.18

    # Calculate water amount
    water_amount = quantity * WATER_RATE

    # Calculate transportation amount
    transport_amount = quantity * TRANSPORT_RATE

    # Calculate subtotal
    subtotal = water_amount + transport_amount

    # Calculate GST
    gst = subtotal * GST_RATE

    # Calculate final amount
    total = subtotal + gst

    info = {
        "order_id": order.get("order_id"),
        "stp_id": order.get("stp_id"),
        "stp_name": order.get("stp_name"),

        "quantity": order.get("quantity_kld"),
        "quality_required": order.get("quality"),
        "water_type": order.get("water_type"),

        "distance_km": order.get("distance_km"),
        "location": order.get("location"),

        "buyer_name": order.get("buyer_name"),
        "buyer_phone": order.get("buyer_phone"),

        "status": order.get("status"),
        "created_at": order.get("created_at"),

        # Invoice amounts
        "water_rate": f"{WATER_RATE:.2f}",
        "water_amount": f"{water_amount:.2f}",
        "transport_rate": f"{TRANSPORT_RATE:.2f}",
        "transport_amount": f"{transport_amount:.2f}",
        "subtotal": f"{subtotal:.2f}",
        "gst": f"{gst:.2f}",
        "total": f"{total:.2f}"
    }

    return render_template(
        "invoice.html",
        info=info,
        invoice_date=order.get("created_at")
    )


# =========================================================
# PAYMENT / BOOKING CONFIRMATION
# =========================================================

@app.route("/pay_now", methods=["POST"])
def pay_now():
    order_id = request.form.get("order_id", "").strip()

    if not order_id:
        return "Order ID is missing", 400

    if not os.path.exists(ORDERS_FILE):
        return "Orders file not found", 404

    updated_rows = []
    order_found = False

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:

            # Remove unnamed CSV columns
            row.pop(None, None)

            if row.get("order_id", "").strip() == order_id:
                current_user_id = session.get("user_id")
                current_buyer_name = session.get("buyer_name") or session.get("user_name")
                current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

                authorized = (
                    current_user_id and
                    row.get("buyer_user_id", "") == current_user_id
                ) or (
                    not row.get("buyer_user_id", "") and
                    current_buyer_name and
                    current_buyer_phone and
                    row.get("buyer_name") == current_buyer_name and
                    row.get("buyer_phone") == current_buyer_phone
                )

                if not authorized:
                    return "Unauthorized", 403

                row["status"] = "Pending"
                row["payment_status"] = "Paid"
                order_found = True

            updated_rows.append(row)

    if not order_found:
        return "Order not found", 404

    # Keep every existing column and add payment_status when needed.
    if "payment_status" not in fieldnames:
        fieldnames.append("payment_status")

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)

    return redirect(url_for("demand", payment_success=order_id))


@app.route("/confirm_cod", methods=["POST"])
def confirm_cod():
    order_id = request.form.get("order_id", "").strip()

    if not order_id:
        return "Order ID is missing", 400

    if not os.path.exists(ORDERS_FILE):
        return "Orders file not found", 404

    updated_rows = []
    order_found = False

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:
            row.pop(None, None)
            if row.get("order_id", "").strip() == order_id:
                current_user_id = session.get("user_id")
                current_buyer_name = session.get("buyer_name") or session.get("user_name")
                current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

                authorized = (
                    current_user_id and
                    row.get("buyer_user_id", "") == current_user_id
                ) or (
                    not row.get("buyer_user_id", "") and
                    current_buyer_name and
                    current_buyer_phone and
                    row.get("buyer_name") == current_buyer_name and
                    row.get("buyer_phone") == current_buyer_phone
                )

                if not authorized:
                    return "Unauthorized", 403

                row["status"] = "Pending"
                row["payment_status"] = "Cash on Delivery"
                order_found = True

            updated_rows.append(row)

    if not order_found:
        return "Order not found", 404

    # Keep every existing column and add payment_status when needed.
    if "payment_status" not in fieldnames:
        fieldnames.append("payment_status")

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)

    return redirect(url_for("demand", payment_success=order_id))




@app.route("/api/my_orders")
def my_orders():
    user_id = session.get("user_id")
    buyer_name = session.get("buyer_name") or session.get("user_name")
    buyer_phone = session.get("buyer_phone") or session.get("user_phone")

    if not user_id and not buyer_name and not buyer_phone:
        return jsonify({"error": "Please log in to view your orders."}), 401

    results = []
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                matches_user = bool(user_id and row.get("buyer_user_id", "") == user_id)
                matches_legacy = (
                    not row.get("buyer_user_id", "") and buyer_name and buyer_phone and
                    row.get("buyer_name") == buyer_name and row.get("buyer_phone") == buyer_phone
                )
                if matches_user or matches_legacy:
                    results.append({
                        "order_id": row.get("order_id"),
                        "status": row.get("status"),
                        "location": row.get("location"),
                        "stp_name": row.get("stp_name"),
                        "quantity_kld": row.get("quantity_kld"),
                        "created_at": row.get("created_at"),
                        "payment_status": row.get("payment_status", "")
                    })

    results.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return jsonify(results)

@app.route("/api/order_tracking/<order_id>")
def order_tracking(order_id):

    if not os.path.exists(ORDERS_FILE):
        return jsonify({
            "success": False,
            "error": "Orders file not found"
        }), 404

    order = None

    with open(
        ORDERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:
            if str(row.get("order_id", "")).strip() == str(order_id).strip():
                order = row
                break

    if not order:
        return jsonify({
            "success": False,
            "error": "Order not found"
        }), 404

    # -----------------------------------------
    # FIND STP
    # -----------------------------------------

    stps = load_stps()

    stp = None

    for s in stps:
        if str(s.get("stp_id", "")).strip() == str(order.get("stp_id", "")).strip():
            stp = s
            break

    if not stp:
        return jsonify({
            "success": False,
            "error": "STP not found"
        }), 404

    # -----------------------------------------
    # GEOCODE DELIVERY LOCATION
    # -----------------------------------------

    location = str(order.get("location", "")).strip()

    delivery_lat = None
    delivery_lon = None

    if location:

        try:
            geo_url = (
                "https://nominatim.openstreetmap.org/search"
                "?format=json"
                "&limit=1"
                "&countrycodes=in"
                "&q="
                + requests.utils.quote(
                    location + ", Bangalore"
                )
            )

            response = requests.get(
                geo_url,
                headers={
                    "User-Agent": "wastewater-app"
                },
                timeout=10
            )

            geo_data = response.json()

            if geo_data:
                delivery_lat = float(geo_data[0]["lat"])
                delivery_lon = float(geo_data[0]["lon"])

        except Exception as e:
            print(
                "Tracking geocoding error:",
                e
            )

    # -----------------------------------------
    # RETURN COMPLETE TRACKING DATA
    # -----------------------------------------

    return jsonify({
        "success": True,

        "order_id": order.get("order_id", ""),

        "status": order.get(
            "status",
            "Pending"
        ),

        "stp": {
            "id": stp.get("stp_id", ""),
            "name": stp.get("stp_name", ""),
            "latitude": float(stp.get("latitude")),
            "longitude": float(stp.get("longitude"))
        },

        "delivery": {
            "location": location,
            "latitude": delivery_lat,
            "longitude": delivery_lon
        }
    })

@app.route("/api/track_order")
def track_order():

    order_id = request.args.get("order_id")
    phone = request.args.get("phone")
    
    if not order_id and not phone:
        return jsonify({"error": "Provide order_id or phone"}), 400

    results = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)

            for row in reader:
                if (
                    (order_id and row.get("order_id") == order_id) or
                    (phone and row.get("buyer_phone") == phone)
                ):
                    results.append({
                        "order_id": row.get("order_id"),
                        "status": row.get("status"),
                        "location": row.get("location"),
                        "stp_name": row.get("stp_name"),
                        "created_at": row.get("created_at")
                    })

    results.sort(key=lambda x: x["order_id"], reverse=True)
    return jsonify(results)

# =========================================================
# SUPPLY SIDE
# =========================================================

@app.route('/supply')
def supply():


    auto_reset_capacity()

    stps = load_stps()
    selected_id = request.args.get("stp_id")
    selected_stp = None
    prediction = None
    weekly_forecast = None

    if selected_id:
        for stp in stps:
            if str(stp["stp_id"]) == str(selected_id):
                selected_stp = stp

                try:
                    print("STP ID sent to ML:", stp["stp_id"])

                    prediction = predict_next_day(str(stp["stp_id"]))
                    weekly_forecast = predict_week(str(stp["stp_id"]))

                    if prediction is not None:
                        prediction = round(prediction, 2)

                    print("Prediction:", prediction)
                except Exception as e:
                    print("Prediction error:", e)
                    prediction = None

    demands = []

    # =========================================================
    # STP-TO-STP TRANSFER REQUESTS
    # =========================================================

    transfer_requests = []

    if selected_stp and os.path.exists(STP_TRANSFERS_FILE):

        with open(
            STP_TRANSFERS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)

            for row in reader:

                # This STP is the SOURCE,
                # meaning another STP is requesting water from it.
                if (
                    row.get("source_stp_id", "").strip()
                    == str(selected_stp["stp_id"]).strip()
                ):

                    transfer_requests.append(row)


    # Newest requests first
    transfer_requests.reverse()

    if selected_stp and os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                
                print("ROW STP:", row.get("stp_id"))
                print("SELECTED STP:", selected_stp["stp_id"])

                if row.get("stp_id", "").strip() == str(selected_stp["stp_id"]).strip():
                    
                    print("MATCHED:", row)

                    # ✅ SAFE CLEANING (handles None keys)
                    clean_row = {}

                    for k, v in row.items():
                        if k is None:
                            continue
                        clean_row[k.strip()] = v

                    row = clean_row
                    
                    print("ROW DATA:", row)
                    mapped_row = {
                        "request_id": row.get("order_id"),
                        "site_name": row.get("location"),
                        "quantity": row.get("quantity_kld"),
                        "quality_required": row.get("quality"),
                        "buyer_name": row.get("buyer_name"),        # ✅ ADD THIS
                        "buyer_phone": row.get("buyer_phone"),      # ✅ ADD THIS
                        "status": (row.get("status") or "").strip(),
                        "created_at": row.get("created_at")
                    }

                    demands.append(mapped_row)

    return render_template(
    "supply.html",
    stps=stps,
    selected_stp=selected_stp,
    demands=demands,
    transfer_requests=transfer_requests,
    prediction=prediction,
    weekly_forecast=weekly_forecast
    )

# =========================================================
# REQUEST WATER - STP TO STP
# =========================================================

@app.route('/request-water')
def request_water():

    stps = load_stps()

    selected_id = request.args.get("stp_id")
    selected_stp = None

    if selected_id:
        for stp in stps:
            if str(stp.get("stp_id")) == str(selected_id):
                selected_stp = stp
                break

    # If no valid STP was selected, return to dashboard
    if not selected_stp:
        return redirect(url_for('supply'))

    # Only other STPs can be selected as the source.
    source_stps = [
        stp for stp in stps
        if str(stp.get("stp_id")) != str(selected_stp.get("stp_id"))
    ]

    return render_template(
        "request_water.html",
        selected_stp=selected_stp,
        source_stps=source_stps
    )

@app.route('/request-water/create', methods=['POST'])
def create_stp_transfer():

    data = request.json or {}

    # =========================================================
    # REQUIRED FIELDS
    # =========================================================

    required_fields = [
        "source_stp_id",
        "destination_stp_id",
        "quantity_kld",
        "quality",
        "water_type"
    ]

    missing = [
        field
        for field in required_fields
        if not data.get(field)
    ]

    if missing:
        return jsonify({
            "success": False,
            "error": "Missing required fields",
            "fields": missing
        }), 400


    # =========================================================
    # LOAD STPs
    # =========================================================

    stps = load_stps()

    source_stp = None
    destination_stp = None

    for stp in stps:

        if str(stp.get("stp_id")) == str(
            data["source_stp_id"]
        ):
            source_stp = stp

        if str(stp.get("stp_id")) == str(
            data["destination_stp_id"]
        ):
            destination_stp = stp


    if source_stp is None:

        return jsonify({
            "success": False,
            "error": "Source STP not found"
        }), 404


    if destination_stp is None:

        return jsonify({
            "success": False,
            "error": "Destination STP not found"
        }), 404


    # =========================================================
    # SOURCE AND DESTINATION MUST BE DIFFERENT
    # =========================================================

    if (
        str(source_stp["stp_id"])
        == str(destination_stp["stp_id"])
    ):

        return jsonify({
            "success": False,
            "error": "Source and destination STP cannot be the same"
        }), 400


    # =========================================================
    # VALIDATE QUANTITY
    # =========================================================

    try:

        quantity_kld = float(
            data["quantity_kld"]
        )

    except (TypeError, ValueError):

        return jsonify({
            "success": False,
            "error": "Invalid quantity"
        }), 400


    if quantity_kld <= 0:

        return jsonify({
            "success": False,
            "error": "Quantity must be greater than zero"
        }), 400


    # =========================================================
    # SOURCE AVAILABLE CAPACITY
    #
    # STP dataset = MLD
    # Request = KLD
    # =========================================================

    try:

        available_mld = float(
            source_stp.get(
                "available_capacity_mld",
                0
            ) or 0
        )

    except (TypeError, ValueError):

        available_mld = 0.0


    available_kld = available_mld * 1000


    if quantity_kld > available_kld:

        return jsonify({
            "success": False,
            "error": (
                "Requested quantity exceeds "
                "available source STP capacity"
            ),
            "available_kld": round(
                available_kld,
                2
            )
        }), 400


    # =========================================================
    # QUALITY VALIDATION
    # =========================================================

    requested_quality = (
        str(data["quality"]).strip()
    )

    source_quality = (
        str(
            source_stp.get(
                "quality_grade",
                ""
            )
        ).strip()
    )


    if (
        requested_quality
        and source_quality
        and requested_quality.lower()
        != source_quality.lower()
    ):

        return jsonify({
            "success": False,
            "error": (
                "Requested water quality is "
                "not available at the source STP"
            ),
            "source_quality": source_quality
        }), 400


    # =========================================================
    # WATER TYPE VALIDATION
    # =========================================================

    requested_type = (
        str(data["water_type"]).strip()
    )

    source_type = (
        str(
            source_stp.get(
                "water_type",
                ""
            )
        ).strip()
    )


    if (
        requested_type
        and source_type
        and requested_type.lower()
        != source_type.lower()
    ):

        return jsonify({
            "success": False,
            "error": (
                "Requested water type is "
                "not supported by the source STP"
            ),
            "source_water_type": source_type
        }), 400


    # =========================================================
    # DISTANCE
    # =========================================================

    distance_km = astar_distance(
        float(source_stp["latitude"]),
        float(source_stp["longitude"]),
        float(destination_stp["latitude"]),
        float(destination_stp["longitude"])
    )


    # =========================================================
    # CREATE TRANSFER ID
    # =========================================================

    transfer_id = (
        "TRF-"
        + uuid.uuid4().hex[:8].upper()
    )


    # =========================================================
    # CREATE RECORD
    # =========================================================

    row = {

        "transfer_id": transfer_id,

        "source_stp_id":
            source_stp["stp_id"],

        "source_stp_name":
            source_stp["stp_name"],

        "destination_stp_id":
            destination_stp["stp_id"],

        "destination_stp_name":
            destination_stp["stp_name"],

        "quantity_kld":
            quantity_kld,

        "quality":
            requested_quality,

        "water_type":
            requested_type,

        "distance_km":
            round(distance_km, 2),

        "status":
            "Pending",

        "requested_at":
            datetime.now().isoformat(),

        "accepted_at":
            "",

        "rejected_at":
            "",

        "tanker_status":
            "Not Assigned"
    }


    # =========================================================
    # SAVE REQUEST
    # =========================================================

    with open(
        STP_TRANSFERS_FILE,
        "a",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=STP_TRANSFER_FIELDS
        )

        writer.writerow(row)


    # =========================================================
    # RESPONSE
    # =========================================================

    return jsonify({

        "success": True,

        "message":
            "Water transfer request submitted successfully",

        "transfer_id":
            transfer_id,

        "distance_km":
            round(distance_km, 2),

        "status":
            "Pending"
    })

# =========================================================
# HANDLE STP-TO-STP TRANSFER REQUEST
# =========================================================

@app.route("/handle_transfer_request", methods=["POST"])
def handle_transfer_request():

    transfer_id = (request.form.get("transfer_id") or "").strip()
    action = (request.form.get("action") or "").strip().lower()

    if not transfer_id:
        return "Transfer ID is required", 400

    if action not in {"accept", "reject"}:
        return "Invalid action", 400

    ensure_stp_transfers_file()

    updated_rows = []
    source_stp_id = None
    found = False

    # ---------------------------------------------------------
    # READ TRANSFER REQUESTS
    # ---------------------------------------------------------

    with open(
        STP_TRANSFERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            if row.get("transfer_id", "").strip() != transfer_id:
                updated_rows.append(row)
                continue

            found = True

            source_stp_id = row.get("source_stp_id")

            current_status = (
                row.get("status") or ""
            ).strip()

            # Only pending requests can be accepted/rejected
            if current_status != "Pending":
                updated_rows.append(row)
                continue

            # =================================================
            # REJECT
            # =================================================

            if action == "reject":

                row["status"] = "Rejected"

                row["rejected_at"] = (
                    datetime.now().isoformat()
                )

                updated_rows.append(row)

                continue

            # =================================================
            # ACCEPT
            # =================================================

            stps = load_stps()

            source_stp = None

            for stp in stps:

                if str(stp.get("stp_id")) == str(source_stp_id):

                    source_stp = stp
                    break

            if source_stp is None:
                return "Source STP not found", 404

            # -------------------------------------------------
            # QUANTITY
            # -------------------------------------------------

            try:

                quantity_kld = float(
                    row.get("quantity_kld") or 0
                )

            except (TypeError, ValueError):

                return "Invalid transfer quantity", 400

            if quantity_kld <= 0:
                return "Transfer quantity must be greater than zero", 400

            # KLD → MLD
            quantity_mld = quantity_kld / 1000.0

            # -------------------------------------------------
            # CHECK CAPACITY
            # -------------------------------------------------

            try:

                available_mld = float(
                    source_stp.get(
                        "available_capacity_mld",
                        0
                    ) or 0
                )

            except (TypeError, ValueError):

                available_mld = 0.0

            if available_mld < quantity_mld:

                return (
                    "Insufficient STP capacity",
                    400
                )

            # -------------------------------------------------
            # RESERVE WATER
            # -------------------------------------------------

            source_stp["available_capacity_mld"] = round(
                available_mld - quantity_mld,
                6
            )

            source_stp["current_load_mld"] = round(
                float(
                    source_stp.get(
                        "current_load_mld",
                        0
                    ) or 0
                ) + quantity_mld,
                6
            )

            save_stps(stps)

            # -------------------------------------------------
            # UPDATE REQUEST
            # -------------------------------------------------

            row["status"] = "Accepted"

            row["accepted_at"] = (
                datetime.now().isoformat()
            )

            row["rejected_at"] = ""

            row["tanker_status"] = (
                "Pending Assignment"
            )

            updated_rows.append(row)

    # =========================================================
    # REQUEST NOT FOUND
    # =========================================================

    if not found:
        return "Transfer request not found", 404

    # =========================================================
    # SAVE UPDATED CSV
    # =========================================================

    with open(
        STP_TRANSFERS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=STP_TRANSFER_FIELDS
        )

        writer.writeheader()
        writer.writerows(updated_rows)

import os
# =========================================================
# WASTEWATER CHATBOT API
# =========================================================

@app.route("/api/chat", methods=["POST"])
def chatbot():
    """
    Wastewater Assistant API.

    Supported buyer-facing intents:
    - greetings / help
    - current account role
    - STP count and availability
    - nearest STP using browser/session location
    - suitable STP recommendation using required KLD
    - latest / previous order
    - complete order history
    - total ordered quantity / order count
    - latest order status
    - tanker status
    - delivery status
    - order lookup by order ID

    The chatbot reads the same STP and orders data used by the rest of
    the application, so it does not maintain a separate chatbot database.
    """
    import re
    import traceback

    try:
        data = request.get_json(silent=True) or {}

        message = str(data.get("message") or "").strip()
        if not message:
            return jsonify({"reply": "Please type a question."}), 400

        text = re.sub(r"\s+", " ", message.lower()).strip()

        # ---------------------------------------------------------
        # LOCATION
        # ---------------------------------------------------------
        latitude = data.get("latitude")
        longitude = data.get("longitude")

        try:
            latitude = float(latitude) if latitude not in (None, "") else None
            longitude = float(longitude) if longitude not in (None, "") else None
        except (TypeError, ValueError):
            latitude = None
            longitude = None

        # If the browser did not send a location, reuse the exact location
        # saved by the Demand search page.
        if latitude is None or longitude is None:
            saved_location = session.get("last_demand_location") or {}
            try:
                if latitude is None and saved_location.get("latitude") is not None:
                    latitude = float(saved_location["latitude"])
                if longitude is None and saved_location.get("longitude") is not None:
                    longitude = float(saved_location["longitude"])
            except (TypeError, ValueError):
                pass

        # ---------------------------------------------------------
        # CURRENT USER
        # ---------------------------------------------------------
        role = str(session.get("role") or "guest").strip().lower()
        user_id = str(session.get("user_id") or "").strip()
        buyer_name = str(
            session.get("buyer_name")
            or session.get("user_name")
            or ""
        ).strip()
        buyer_phone = str(
            session.get("buyer_phone")
            or session.get("user_phone")
            or ""
        ).strip()

        # ---------------------------------------------------------
        # QUANTITY EXTRACTION
        # ---------------------------------------------------------
        requested_kld = None

        quantity_match = re.search(
            r"(\d+(?:\.\d+)?)\s*(kld|kl|litres?|liters?)\b",
            text
        )

        if quantity_match:
            quantity_value = float(quantity_match.group(1))
            unit = quantity_match.group(2).lower()

            if unit in {"litre", "litres", "liter", "liters"}:
                requested_kld = quantity_value / 1000.0
            else:
                requested_kld = quantity_value

        # ---------------------------------------------------------
        # COMMON INTENTS
        # ---------------------------------------------------------
        greetings = {
            "hi",
            "hello",
            "hey",
            "hai",
            "good morning",
            "good afternoon",
            "good evening",
        }

        if text in greetings:
            return jsonify({
                "reply": (
                    "Hello! 👋 I'm your Wastewater Assistant.\n\n"
                    "I can help you with STPs, orders, routing, "
                    "demand, predictions and tanker information."
                )
            })

        if (
            "what can you do" in text
            or "what do you do" in text
            or text in {"help", "help me"}
        ):
            return jsonify({
                "reply": (
                    "I can help with:\n\n"
                    "🏭 STP locations and availability\n"
                    "📦 Latest order and order history\n"
                    "📌 Order status\n"
                    "💧 Ordered quantity and totals\n"
                    "🚚 Tanker and delivery status\n"
                    "📍 Nearest STP\n"
                    "🎯 Suitable STP recommendations\n"
                    "🗺️ Routing information\n"
                    "📈 Demand and prediction information"
                )
            })

        if (
            "my role" in text
            or "who am i" in text
            or "my account" in text
        ):
            if role == "guest":
                return jsonify({
                    "reply": "You are currently not logged in."
                })

            role_names = {
                "demand": "Site User / Buyer",
                "stp": "STP / Seller",
                "tanker": "Tanker Operator",
                "admin": "Administrator",
            }

            return jsonify({
                "reply": (
                    f"You are logged in as "
                    f"{role_names.get(role, role.title())}."
                )
            })

        # ---------------------------------------------------------
        # STP INFORMATION
        # ---------------------------------------------------------
        stp_info_query = (
            "stp" in text
            and any(
                phrase in text
                for phrase in (
                    "how many",
                    "number",
                    "available",
                    "list",
                    "show",
                    "all stp",
                    "all the stp",
                )
            )
        )

        if stp_info_query:
            stps = load_stps()

            if not stps:
                return jsonify({
                    "reply": "There are currently no STPs available in the system."
                })

            available = []
            for stp in stps:
                try:
                    capacity_mld = float(
                        stp.get("available_capacity_mld") or 0
                    )
                except (TypeError, ValueError):
                    capacity_mld = 0.0

                if capacity_mld > 0:
                    available.append((stp, capacity_mld))

            reply_lines = [
                f"🏭 There are {len(stps)} STPs in the system.",
                f"💧 {len(available)} currently have available capacity.",
            ]

            if available:
                reply_lines.append("")
                reply_lines.append("Available STPs:")
                for stp, capacity_mld in available[:10]:
                    name = (
                        stp.get("stp_name")
                        or stp.get("name")
                        or stp.get("stp_id")
                        or "Unnamed STP"
                    )
                    reply_lines.append(
                        f"• {name} — {capacity_mld * 1000:.0f} KLD available"
                    )

                if len(available) > 10:
                    reply_lines.append(
                        f"• ...and {len(available) - 10} more."
                    )

            return jsonify({"reply": "\n".join(reply_lines)})

        # ---------------------------------------------------------
        # NEAREST STP
        # ---------------------------------------------------------
        nearest_stp_query = any(
            phrase in text
            for phrase in (
                "nearest stp",
                "closest stp",
                "stp near me",
                "stp nearby",
                "nearest stp to me",
                "closest stp to me",
                "which stp is near",
                "which stp is closest",
                "where is the nearest stp",
                "where is the closest stp",
                "what is the nearest stp",
                "what's the nearest stp",
                "find the nearest stp",
                "find the closest stp",
            )
        )

        if nearest_stp_query:
            if latitude is None or longitude is None:
                return jsonify({
                    "reply": (
                        "📍 I need your location to find the nearest STP.\n\n"
                        "Please allow location access in your browser and "
                        "try again."
                    )
                })

            stps = load_stps()
            if not stps:
                return jsonify({
                    "reply": "I couldn't find any STPs in the system."
                })

            nearest_stp = None
            nearest_distance = float("inf")

            for stp in stps:
                try:
                    stp_lat = float(stp.get("latitude"))
                    stp_lon = float(stp.get("longitude"))
                except (TypeError, ValueError):
                    continue

                distance = haversine(
                    latitude,
                    longitude,
                    stp_lat,
                    stp_lon
                )

                if distance < nearest_distance:
                    nearest_distance = distance
                    nearest_stp = stp

            if nearest_stp is None:
                return jsonify({
                    "reply": (
                        "I found STPs in the system, but their "
                        "location coordinates are unavailable."
                    )
                })

            stp_name = (
                nearest_stp.get("stp_name")
                or nearest_stp.get("name")
                or nearest_stp.get("stp_id")
                or "Nearest STP"
            )

            try:
                available_kld = (
                    float(nearest_stp.get("available_capacity_mld") or 0)
                    * 1000
                )
                capacity_text = f"{available_kld:.0f} KLD"
            except (TypeError, ValueError):
                capacity_text = "Unknown"

            return jsonify({
                "reply": (
                    "📍 Nearest STP\n\n"
                    f"🏭 STP: {stp_name}\n"
                    f"📏 Distance: {nearest_distance:.2f} km\n"
                    f"💧 Available Capacity: {capacity_text}"
                )
            })

        # ---------------------------------------------------------
        # SMART STP RECOMMENDATION
        # ---------------------------------------------------------
        recommendation_query = any(
            phrase in text
            for phrase in (
                "which stp should i choose",
                "which stp should i select",
                "which stp is best",
                "recommend an stp",
                "recommend a stp",
                "find an stp",
                "suitable stp",
                "best stp",
                "stp for me",
                "stp for my requirement",
                "need an stp",
                "which stp can provide",
                "where can i get",
            )
        )

        if recommendation_query:
            if requested_kld is None:
                return jsonify({
                    "reply": (
                        "🎯 I can recommend a suitable STP.\n\n"
                        "Please tell me the required quantity, for example:\n"
                        "“Which STP is suitable for 20 KLD?”"
                    )
                })

            if requested_kld <= 0:
                return jsonify({
                    "reply": "Please provide a quantity greater than 0 KLD."
                })

            if latitude is None or longitude is None:
                return jsonify({
                    "reply": (
                        "📍 I need your location to recommend the nearest "
                        "suitable STP. Please allow location access and try again."
                    )
                })

            suitable_stps = []

            for stp in load_stps():
                try:
                    available_mld = float(
                        stp.get("available_capacity_mld") or 0
                    )
                    available_kld = available_mld * 1000

                    stp_lat = float(stp.get("latitude"))
                    stp_lon = float(stp.get("longitude"))
                except (TypeError, ValueError):
                    continue

                if available_kld < requested_kld:
                    continue

                distance = haversine(
                    latitude,
                    longitude,
                    stp_lat,
                    stp_lon
                )

                stp_name = (
                    stp.get("stp_name")
                    or stp.get("name")
                    or stp.get("stp_id")
                    or "Unnamed STP"
                )

                suitable_stps.append({
                    "name": stp_name,
                    "stp_id": stp.get("stp_id", ""),
                    "distance": distance,
                    "available_kld": available_kld,
                    "quality": stp.get("quality_grade") or "Unknown",
                    "water_type": stp.get("water_type") or "Unknown",
                })

            if not suitable_stps:
                return jsonify({
                    "reply": (
                        f"🎯 I couldn't find an STP near you with at least "
                        f"{requested_kld:g} KLD of available capacity."
                    )
                })

            suitable_stps.sort(key=lambda item: item["distance"])
            top_stps = suitable_stps[:3]
            best = top_stps[0]

            reply = (
                f"🎯 I found {len(suitable_stps)} suitable STP(s) "
                f"for {requested_kld:g} KLD.\n\n"
                f"🏆 Recommended STP\n\n"
                f"🏭 STP: {best['name']}\n"
                f"📏 Distance: {best['distance']:.2f} km\n"
                f"💧 Available Capacity: {best['available_kld']:.0f} KLD\n"
            )

            if str(best["quality"]).strip().lower() != "unknown":
                reply += f"🧪 Quality: {best['quality']}\n"

            if str(best["water_type"]).strip().lower() != "unknown":
                reply += f"💦 Water Type: {best['water_type']}\n"

            if len(top_stps) > 1:
                reply += "\nOther suitable options:\n"
                for index, stp in enumerate(top_stps[1:], start=2):
                    reply += (
                        f"{index}. {stp['name']} — "
                        f"{stp['distance']:.2f} km away, "
                        f"{stp['available_kld']:.0f} KLD available\n"
                    )

            return jsonify({"reply": reply})

        # ---------------------------------------------------------
        # ORDER INTENTS
        # ---------------------------------------------------------
        history_query = any(
            phrase in text
            for phrase in (
                "order history",
                "my order history",
                "show my orders",
                "show my order history",
                "what orders have i placed",
                "what orders did i place",
                "orders have i placed",
                "orders did i place",
                "previous orders",
                "all my orders",
            )
        )

        latest_order_query = any(
            phrase in text
            for phrase in (
                "previous order",
                "what was my previous order",
                "last order",
                "latest order",
                "recent order",
                "what did i order last",
                "what was my last order",
                "what is my previous order",
                "what is my latest order",
            )
        )

        total_quantity_query = any(
            phrase in text
            for phrase in (
                "total water",
                "total quantity",
                "total kld",
                "how much water have i ordered",
                "how much have i ordered",
                "how much water did i order in total",
                "total amount of water",
            )
        )

        order_count_query = any(
            phrase in text
            for phrase in (
                "how many orders have i made",
                "how many orders did i make",
                "how many orders have i placed",
                "number of orders i placed",
                "how many orders do i have",
            )
        )

        quantity_query = any(
            phrase in text
            for phrase in (
                "how much water did i order",
                "how much did i order",
                "what quantity did i order",
                "how many kld did i order",
                "what is my order quantity",
            )
        )

        status_query = (
            "order status" in text
            or "status of my order" in text
            or "what's my order status" in text
            or "what is my order status" in text
            or "whats my order status" in text
            or "what is the order status" in text
            or "what's the order status" in text
            or "whats the order status" in text
            or "check my order" in text
            or "track my order" in text
        )

        tanker_query = any(
            phrase in text
            for phrase in (
                "where is my tanker",
                "tanker status",
                "has my tanker been assigned",
                "is my tanker assigned",
                "tanker assigned",
            )
        )

        delivery_query = any(
            phrase in text
            for phrase in (
                "delivery status",
                "what is my delivery status",
                "what's my delivery status",
                "whats my delivery status",
                "where is my delivery",
                "when will my delivery arrive",
                "when will my order arrive",
            )
        )

        order_id_match = re.search(
            r"\bORD-[A-Z0-9]+\b",
            message,
            flags=re.IGNORECASE
        )
        requested_order_id = (
            order_id_match.group(0).upper()
            if order_id_match
            else None
        )

        order_related = (
            history_query
            or latest_order_query
            or total_quantity_query
            or order_count_query
            or quantity_query
            or status_query
            or tanker_query
            or delivery_query
            or requested_order_id is not None
        )

        if order_related:
            if not user_id:
                return jsonify({
                    "reply": (
                        "🔐 Please log in first so I can securely "
                        "access your orders."
                    )
                })

            orders = []

            if os.path.exists(ORDERS_FILE):
                with open(
                    ORDERS_FILE,
                    "r",
                    newline="",
                    encoding="utf-8-sig"
                ) as f:
                    reader = csv.DictReader(f)

                    for raw_row in reader:
                        row = {
                            str(key).strip(): (value or "").strip()
                            for key, value in raw_row.items()
                            if key is not None
                        }

                        row_user_id = str(
                            row.get("buyer_user_id") or ""
                        ).strip()

                        row_name = str(
                            row.get("buyer_name") or ""
                        ).strip()

                        row_phone = str(
                            row.get("buyer_phone") or ""
                        ).strip()

                        matches_user = (
                            bool(user_id)
                            and bool(row_user_id)
                            and row_user_id == user_id
                        )

                        # Backward compatibility for orders created before
                        # buyer_user_id was added.
                        matches_legacy = (
                            not row_user_id
                            and bool(buyer_name)
                            and bool(buyer_phone)
                            and row_name == buyer_name
                            and row_phone == buyer_phone
                        )

                        if matches_user or matches_legacy:
                            orders.append(row)

            if requested_order_id:
                orders = [
                    row
                    for row in orders
                    if str(row.get("order_id") or "").strip().upper()
                    == requested_order_id
                ]

            if not orders:
                if requested_order_id:
                    return jsonify({
                        "reply": (
                            f"I couldn't find order {requested_order_id} "
                            f"associated with your account."
                        )
                    })

                return jsonify({
                    "reply": (
                        "I couldn't find any orders associated "
                        "with your account."
                    )
                })

            orders.sort(
                key=lambda row: row.get("created_at") or "",
                reverse=True
            )

            # -----------------------------------------------------
            # COMPLETE HISTORY
            # -----------------------------------------------------
            if history_query and not latest_order_query:
                history_lines = ["📦 Order History", ""]

                for index, order in enumerate(orders, start=1):
                    order_id = order.get("order_id") or "Unknown"
                    quantity = order.get("quantity_kld") or "Unknown"
                    stp_name = (
                        order.get("stp_name")
                        or order.get("stp_id")
                        or "Unknown STP"
                    )
                    status = order.get("status") or "Unknown"

                    history_lines.append(
                        f"{index}. {order_id}\n"
                        f"   💧 Quantity: {quantity} KLD\n"
                        f"   🏭 STP: {stp_name}\n"
                        f"   📌 Status: {status}"
                    )

                history_lines.append("")
                history_lines.append(
                    f"You have placed {len(orders)} order(s)."
                )

                return jsonify({
                    "reply": "\n\n".join(history_lines)
                })

            latest = orders[0]

            order_id = latest.get("order_id") or "Unknown"
            quantity = latest.get("quantity_kld") or "Unknown"
            stp_name = (
                latest.get("stp_name")
                or latest.get("stp_id")
                or "Unknown STP"
            )
            status = latest.get("status") or "Unknown"
            location = latest.get("location") or "your delivery location"
            payment_status = latest.get("payment_status") or "Unknown"
            created_at = latest.get("created_at") or "Unknown"

            # -----------------------------------------------------
            # TOTAL QUANTITY
            # -----------------------------------------------------
            if total_quantity_query:
                total_kld = 0.0

                for order in orders:
                    try:
                        total_kld += float(order.get("quantity_kld") or 0)
                    except (TypeError, ValueError):
                        continue

                return jsonify({
                    "reply": (
                        "💧 Total Ordered Quantity\n\n"
                        f"You have ordered {total_kld:g} KLD "
                        f"across {len(orders)} order(s)."
                    )
                })

            # -----------------------------------------------------
            # ORDER COUNT
            # -----------------------------------------------------
            if order_count_query:
                return jsonify({
                    "reply": (
                        f"📦 You have placed {len(orders)} order(s)."
                    )
                })

            # -----------------------------------------------------
            # LATEST / PREVIOUS ORDER DETAILS
            # -----------------------------------------------------
            if latest_order_query or quantity_query:
                if quantity_query and not latest_order_query:
                    return jsonify({
                        "reply": (
                            f"💧 Your latest order {order_id} is for "
                            f"{quantity} KLD of treated wastewater "
                            f"from {stp_name}."
                        )
                    })

                return jsonify({
                    "reply": (
                        "📦 Latest Order\n\n"
                        f"🆔 Order ID: {order_id}\n"
                        f"💧 Quantity: {quantity} KLD\n"
                        f"🏭 STP: {stp_name}\n"
                        f"📌 Status: {status}\n"
                        f"💳 Payment: {payment_status}\n"
                        f"📅 Created: {created_at}"
                    )
                })

            # -----------------------------------------------------
            # STATUS / TANKER / DELIVERY
            # -----------------------------------------------------
            if status_query or tanker_query or delivery_query:
                status_normalized = status.strip().lower()

                if status_query:
                    if status_normalized == "pending":
                        reply = (
                            "📦 Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            "📌 Status: Pending\n"
                            "The order is awaiting STP approval."
                        )
                    elif status_normalized == "accepted":
                        reply = (
                            "📦 Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            "📌 Status: Accepted\n"
                            f"🏭 STP: {stp_name}\n"
                            "The order is waiting for tanker pickup."
                        )
                    elif status_normalized == "out for delivery":
                        reply = (
                            "🚚 Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            "📌 Status: Out for Delivery\n"
                            f"🏭 STP: {stp_name}\n"
                            f"💧 Quantity: {quantity} KLD\n"
                            f"📍 Delivery: {location}"
                        )
                    elif status_normalized == "delivered":
                        reply = (
                            "✅ Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            "📌 Status: Delivered\n"
                            f"🏭 STP: {stp_name}\n"
                            f"💧 Quantity: {quantity} KLD"
                        )
                    elif status_normalized == "rejected":
                        reply = (
                            "❌ Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            "📌 Status: Rejected\n\n"
                            "I can help you find another suitable STP."
                        )
                    else:
                        reply = (
                            "📦 Order Status\n\n"
                            f"🆔 Order: {order_id}\n"
                            f"📌 Status: {status}"
                        )

                    return jsonify({"reply": reply})

                if tanker_query:
                    if status_normalized == "pending":
                        reply = (
                            "🚚 Tanker Status\n\n"
                            f"Order {order_id} is still Pending.\n"
                            "A tanker has not been assigned because "
                            "the order is awaiting STP approval."
                        )
                    elif status_normalized == "accepted":
                        reply = (
                            "🚚 Tanker Status\n\n"
                            f"Order {order_id} has been accepted by "
                            f"{stp_name}.\n"
                            "It is waiting for tanker pickup."
                        )
                    elif status_normalized == "out for delivery":
                        reply = (
                            "🚚 Tanker Status\n\n"
                            f"Order {order_id} is currently Out for Delivery.\n"
                            f"Delivery location: {location}"
                        )
                    elif status_normalized == "delivered":
                        reply = (
                            "✅ Tanker Status\n\n"
                            f"Order {order_id} has already been delivered."
                        )
                    elif status_normalized == "rejected":
                        reply = (
                            "❌ Tanker Status\n\n"
                            f"Order {order_id} was rejected, so a tanker "
                            "has not been assigned."
                        )
                    else:
                        reply = (
                            "🚚 Tanker Status\n\n"
                            f"Order {order_id} currently has status: {status}."
                        )

                    return jsonify({"reply": reply})

                if delivery_query:
                    if status_normalized == "pending":
                        reply = (
                            "📦 Delivery Status\n\n"
                            f"Order {order_id} is still Pending.\n"
                            "Delivery has not started because the order "
                            "is awaiting STP approval."
                        )
                    elif status_normalized == "accepted":
                        reply = (
                            "📦 Delivery Status\n\n"
                            f"Order {order_id} has been accepted by "
                            f"{stp_name}.\n"
                            "It is waiting for tanker pickup."
                        )
                    elif status_normalized == "out for delivery":
                        reply = (
                            "🚚 Delivery Status\n\n"
                            f"Order {order_id} is currently Out for Delivery.\n"
                            f"💧 Quantity: {quantity} KLD\n"
                            f"📍 Delivery: {location}"
                        )
                    elif status_normalized == "delivered":
                        reply = (
                            "✅ Delivery Status\n\n"
                            f"Order {order_id} has been delivered successfully."
                        )
                    elif status_normalized == "rejected":
                        reply = (
                            "❌ Delivery Status\n\n"
                            f"Order {order_id} was rejected, so delivery "
                            "cannot proceed."
                        )
                    else:
                        reply = (
                            "📦 Delivery Status\n\n"
                            f"Order {order_id} currently has status: {status}."
                        )

                    return jsonify({"reply": reply})

        # ---------------------------------------------------------
        # GENERAL SYSTEM GUIDANCE
        # ---------------------------------------------------------
        if "routing" in text or "route" in text:
            return jsonify({
                "reply": (
                    "🗺️ Routing is handled by the application's "
                    "road-network routing module. You can use the "
                    "Routing Map to view routes between the selected "
                    "STP and delivery location."
                )
            })

        if (
            "prediction" in text
            or "forecast" in text
            or "demand prediction" in text
        ):
            return jsonify({
                "reply": (
                    "📈 Demand predictions are available on the STP "
                    "Supply dashboard. Select an STP there to view "
                    "its prediction and weekly forecast."
                )
            })

        if "demand" in text:
            return jsonify({
                "reply": (
                    "💧 Demand information is available through the "
                    "Demand dashboard and its matching STP search. "
                    "Enter your location and required KLD to find "
                    "a suitable treated-wastewater source."
                )
            })

        # ---------------------------------------------------------
        # DEFAULT
        # ---------------------------------------------------------
        return jsonify({
            "reply": (
                "I understood your question, but I don't have a "
                "specific function for it yet.\n\n"
                "Try asking:\n"
                "• “What is my order status?”\n"
                "• “What was my previous order?”\n"
                "• “Show my order history”\n"
                "• “How much water have I ordered?”\n"
                "• “What's the nearest STP?”\n"
                "• “Which STP is suitable for 20 KLD?”"
            )
        })

    except Exception as e:
        print("CHATBOT ERROR:", repr(e))
        traceback.print_exc()

        return jsonify({
            "reply": (
                "Sorry, something went wrong while processing your request. "
                "Please try again."
            )
        }), 500

@app.route("/api/stp_pricing/<stp_id>")
def get_stp_pricing(stp_id):

    pricing = load_stp_pricing()

    for row in pricing:

        if str(row["stp_id"]).strip() == str(stp_id).strip():

            return jsonify({
                "success": True,
                "pricing": {
                    "base_price_per_kld":
                        float(row["base_price_per_kld"]),

                    "peak_incentive":
                        float(row["peak_incentive"]),

                    "off_peak_incentive":
                        float(row["off_peak_incentive"]),

                    "peak_start":
                        row["peak_start"],

                    "peak_end":
                        row["peak_end"],

                    "off_peak_start":
                        row["off_peak_start"],

                    "off_peak_end":
                        row["off_peak_end"],

                    "sustainability_credit":
                        float(row["sustainability_credit"]),

                    "reliability_bonus":
                        float(row["reliability_bonus"])
                }
            })

    return jsonify({
        "success": False,
        "message": "Pricing not found"
    }), 404

@app.route("/api/update_pricing", methods=["POST"])
def update_pricing():

    data = request.get_json()

    if not data:
        return jsonify({
            "success": False,
            "message": "No pricing data received"
        }), 400

    stp_id = data.get("stp_id")

    if not stp_id:
        return jsonify({
            "success": False,
            "message": "STP ID is required"
        }), 400

    try:
        base_price = float(data["base_price_per_kld"])
        peak = float(data["peak_incentive"])
        off_peak = float(data["off_peak_incentive"])
        sustainability = float(data["sustainability_credit"])
        reliability = float(data["reliability_bonus"])

        if base_price < 0:
            raise ValueError

        if peak < 0 or off_peak < 0:
            raise ValueError

        if not 0 <= sustainability <= 100:
            raise ValueError

        if not 0 <= reliability <= 100:
            raise ValueError

    except (ValueError, TypeError, KeyError):

        return jsonify({
            "success": False,
            "message": "Invalid pricing values"
        }), 400

    pricing = load_stp_pricing()
    found = False

    for row in pricing:

        if str(row["stp_id"]).strip() == str(stp_id).strip():

            row["base_price_per_kld"] = base_price
            row["peak_incentive"] = peak
            row["off_peak_incentive"] = off_peak

            row["peak_start"] = data.get("peak_start", "")
            row["peak_end"] = data.get("peak_end", "")

            row["off_peak_start"] = data.get("off_peak_start", "")
            row["off_peak_end"] = data.get("off_peak_end", "")

            row["sustainability_credit"] = sustainability
            row["reliability_bonus"] = reliability

            found = True
            break

    if not found:

        return jsonify({
            "success": False,
            "message": "STP pricing record not found"
        }), 404

    fieldnames = [
        "stp_id",
        "base_price_per_kld",
        "peak_incentive",
        "off_peak_incentive",
        "peak_start",
        "peak_end",
        "off_peak_start",
        "off_peak_end",
        "sustainability_credit",
        "reliability_bonus"
    ]

    with open(
        PRICING_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames
        )

        writer.writeheader()
        writer.writerows(pricing)

    return jsonify({
        "success": True,
        "message": "Pricing updated successfully"
    })

@app.route("/update_capacity", methods=["POST"])
def update_capacity():

    stp_id = request.form["stp_id"]
    new_capacity = float(request.form["available_capacity_mld"])

    stps = load_stps()

    for stp in stps:
        if str(stp["stp_id"]) == str(stp_id):
            stp["available_capacity_mld"] = new_capacity

    save_stps(stps)

    return redirect(url_for("supply", stp_id=stp_id))

@app.route("/upload_quality", methods=["POST"])
def upload_quality():

    stp_id = request.form["stp_id"]
    quality = request.form["quality_grade"]

    stps = load_stps()

    for stp in stps:
        if str(stp["stp_id"]) == str(stp_id):
            stp["quality_grade"] = quality

    save_stps(stps)

    return redirect(url_for("supply", stp_id=stp_id))

@app.route("/handle_request", methods=["POST"])
def handle_request():



    auto_reset_capacity()


    order_id = request.form["request_id"]
    action = request.form.get("action")

    updated_rows = []
    stp_id_redirect = None


    # STEP 1: READ FILE
    if action not in {"accept", "reject"}:
        return "Invalid action", 400

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:

            if row.get("order_id", "").strip() == order_id.strip():

                stp_id_redirect = row.get("stp_id")
                current_status = (row.get("status") or "").strip()

                # Only pending orders can be accepted/rejected by the STP.
                if current_status != "Pending":
                    updated_rows.append(row)
                    continue

                if action == "reject":
                    row["status"] = "Rejected"
                    updated_rows.append(row)
                    continue

                # =====================================================
                # ACCEPT ORDER ONLY IF IT MATCHES THE STP DATASET
                # =====================================================
                stps = load_stps()
                matching_stp = None

                for stp in stps:
                    if str(stp.get("stp_id")) == str(row.get("stp_id")):
                        matching_stp = stp
                        break

                if matching_stp is None:
                    return "STP not found in STP dataset", 404

                try:
                    quantity_kld = float(row.get("quantity_kld") or 0)
                except (TypeError, ValueError):
                    return "Invalid order quantity", 400

                if quantity_kld <= 0:
                    return "Order quantity must be greater than zero", 400

                quantity_mld = quantity_kld / 1000.0

                try:
                    available_capacity = float(
                        matching_stp.get("available_capacity_mld", 0) or 0
                    )
                except (TypeError, ValueError):
                    available_capacity = 0.0

                # Check available STP capacity.
                if available_capacity < quantity_mld:
                    return "Insufficient STP capacity", 400

                # Check requested quality against the STP dataset.
                requested_quality = (row.get("quality") or "").strip()
                stp_quality = (matching_stp.get("quality_grade") or "").strip()

                if (
                    requested_quality
                    and stp_quality
                    and requested_quality.lower() != stp_quality.lower()
                ):
                    return "Requested water quality is not available at this STP", 400

                # Check requested water type against the STP dataset when
                # the STP has a water_type field populated.
                requested_type = (row.get("water_type") or "").strip()
                stp_type = (matching_stp.get("water_type") or "").strip()

                if (
                    requested_type
                    and stp_type
                    and requested_type.lower() != stp_type.lower()
                ):
                    return "Requested water type is not supported by this STP", 400

                # Reserve the requested quantity.
                matching_stp["available_capacity_mld"] = (
                    available_capacity - quantity_mld
                )

                matching_stp["current_load_mld"] = (
                    float(matching_stp.get("current_load_mld", 0) or 0)
                    + quantity_mld
                )

                accepted_at = datetime.now()
                release_at = accepted_at + timedelta(hours=24)

                row["status"] = "Accepted"
                row["accepted_at"] = accepted_at.isoformat()
                row["capacity_release_at"] = release_at.isoformat()
                row["capacity_released"] = "False"

                save_stps(stps)

            updated_rows.append(row)

    if stp_id_redirect is None:
        return "Order not found", 404

    # Keep every existing order column and the new capacity timeline fields.
    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()
        for row in updated_rows:
            writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})

    return redirect(url_for("supply", stp_id=stp_id_redirect))

@app.route("/update_order_status", methods=["POST"])
def update_order_status():
    auto_reset_capacity()

    order_id = (request.form.get("order_id") or "").strip()
    new_status = (request.form.get("status") or "").strip()

    allowed_statuses = {"Pending", "Accepted", "Out for Delivery", "Delivered", "Rejected"}
    if new_status not in allowed_statuses:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    if not order_id:
        return jsonify({"success": False, "error": "Order ID is required"}), 400

    updated = False
    updated_rows = []
    stp_id_redirect = None

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("order_id", "").strip() == order_id:
                stp_id_redirect = row.get("stp_id")
                current_status = row.get("status", "").strip()
                status_order = {"Pending": 0, "Accepted": 1, "Out for Delivery": 2, "Delivered": 3, "Rejected": -1}

                if (
                    current_status != "Rejected" and new_status != "Rejected" and
                    status_order.get(new_status, -1) < status_order.get(current_status, -1)
                ):
                    return jsonify({"success": False, "error": "Cannot move order backwards"}), 400

                if new_status == "Accepted" and current_status != "Accepted":
                    stps = load_stps()
                    quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
                    stp_found = False
                    for stp in stps:
                        if str(stp.get("stp_id")) == str(row.get("stp_id")):
                            available = float(stp.get("available_capacity_mld", 0) or 0)
                            if quantity_mld > available:
                                return jsonify({"success": False, "error": "Insufficient STP capacity"}), 400
                            stp["available_capacity_mld"] = max(0.0, available - quantity_mld)
                            stp["current_load_mld"] = float(stp.get("current_load_mld", 0) or 0) + quantity_mld
                            stp_found = True
                            break
                    if not stp_found:
                        return jsonify({"success": False, "error": "STP not found"}), 404
                    save_stps(stps)
                    accepted_at = datetime.now()
                    row["accepted_at"] = accepted_at.isoformat()
                    row["capacity_release_at"] = (accepted_at + timedelta(hours=24)).isoformat()
                    row["capacity_released"] = "False"

                if (
                    new_status == "Rejected" and
                    current_status in {"Accepted", "Out for Delivery"} and
                    str(row.get("capacity_released", "")).strip().lower() != "true"
                ):
                    stps = load_stps()
                    quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
                    for stp in stps:
                        if str(stp.get("stp_id")) == str(row.get("stp_id")):
                            total = float(stp.get("total_capacity_mld") or 0)
                            available = float(stp.get("available_capacity_mld", 0) or 0)
                            stp["available_capacity_mld"] = min(total, available + quantity_mld)
                            stp["current_load_mld"] = max(0.0, float(stp.get("current_load_mld", 0) or 0) - quantity_mld)
                            row["capacity_released"] = "True"
                            save_stps(stps)
                            break

                row["status"] = new_status
                updated = True
            updated_rows.append(row)

    if not updated:
        return jsonify({"success": False, "error": "Order not found"}), 404

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()
        for row in updated_rows:
            writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})

    return redirect(url_for("supply", stp_id=stp_id_redirect))


@app.route("/trip_history")
def trip_history():
    auto_reset_capacity()

    trip_history = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)

            for row in reader:
                status = (row.get("status") or "").strip()

                if status in {"Accepted", "Out for Delivery", "Delivered"}:
                    trip_history.append(row)

    return render_template("trip_history.html", trip_history=trip_history)
  
TANKER_CAPACITY_KLD = 12
AVAILABLE_TANKERS = 5


@app.route("/tanker")
def tanker_dashboard():

    auto_reset_capacity()

    orders = []

    # =========================================================
    # NORMAL DEMAND ORDERS
    # =========================================================

    if os.path.exists(ORDERS_FILE):

        with open(
            ORDERS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)

            for row in reader:

                if row.get("status") == "Accepted":

                    stps = load_stps()

                    stp_lat = None
                    stp_lon = None

                    for stp in stps:

                        if (
                            str(stp["stp_id"])
                            == str(row["stp_id"])
                        ):

                            stp_lat = stp.get("latitude")
                            stp_lon = stp.get("longitude")

                            break

                    row["stp_lat"] = stp_lat
                    row["stp_lon"] = stp_lon

                    # Mark this as a normal demand order
                    row["request_type"] = "demand"

                    orders.append(row)


    # =========================================================
    # STP → STP TRANSFER REQUESTS
    # =========================================================

    if os.path.exists(STP_TRANSFERS_FILE):

        with open(
            STP_TRANSFERS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)

            for row in reader:

                if (
                    row.get("status", "").strip()
                    in {"Accepted", "Out for Delivery"}
                    and
                    row.get("tanker_status", "").strip()
                    in {"Pending Assignment", "Out for Delivery"}
                ):

                    stps = load_stps()

                    source_stp = None

                    for stp in stps:

                        if (
                            str(stp["stp_id"])
                            == str(row["source_stp_id"])
                        ):

                            source_stp = stp
                            break


                    if source_stp:

                        row["stp_lat"] = source_stp.get(
                            "latitude"
                        )

                        row["stp_lon"] = source_stp.get(
                            "longitude"
                        )

                    else:

                        row["stp_lat"] = None
                        row["stp_lon"] = None


                    # Tell tanker.html what this is
                    row["request_type"] = "stp_transfer"

                    # Fields needed by existing tanker UI
                    row["order_id"] = row.get("transfer_id")

                    row["location"] = row.get(
                        "destination_stp_name"
                    )

                    orders.append(row)


    return render_template(
        "tanker.html",
        orders=orders
    )

@app.route("/accept_pickup", methods=["POST"])
def accept_pickup():

    order_id = request.form.get("order_id")

    if not order_id:
        return "No Order ID received"

    updated_rows = []
    tanker_info = None
    stp_id_redirect = None

    with open(ORDERS_FILE, "r", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames

        for row in reader:

            # Strip spaces to avoid mismatch
            if row["order_id"].strip() == order_id.strip():
                stp_id_redirect = row["stp_id"]

                quantity = float(row["quantity_kld"])

                tankers_required = math.ceil(quantity / TANKER_CAPACITY_KLD)

                tanker_info = {
                    "order_id": row["order_id"],
                    "quantity": quantity,
                    "tankers_required": tankers_required,
                    "available_tankers": AVAILABLE_TANKERS,
                    "sufficient": tankers_required <= AVAILABLE_TANKERS,
                    "buyer_name": row.get("buyer_name"),
                    "buyer_phone": row.get("buyer_phone"),
                }

                row["status"] = "Out for Delivery"

            updated_rows.append(row)

    if tanker_info is None:
        return f"Order {order_id} not found in CSV"

    with open(ORDERS_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)
    
    
    return render_template(
    "tanker_summary.html",
    info=tanker_info,
    stp_id=stp_id_redirect
)

import os

@app.route(
    "/accept_transfer_pickup",
    methods=["POST"]
)
def accept_transfer_pickup():

    transfer_id = (
        request.form.get("transfer_id") or ""
    ).strip()

    if not transfer_id:
        return "No Transfer ID received", 400

    ensure_stp_transfers_file()

    updated_rows = []
    transfer_info = None
    transfer_found = False

    with open(
        STP_TRANSFERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames

        for row in reader:

            if (
                row.get("transfer_id", "").strip()
                == transfer_id
            ):

                transfer_found = True

                if (
                    row.get("status", "").strip()
                    != "Accepted"
                ):

                    updated_rows.append(row)
                    continue

                quantity = float(
                    row.get("quantity_kld", 0) or 0
                )

                tankers_required = math.ceil(
                    quantity / TANKER_CAPACITY_KLD
                )

                transfer_info = {
                    "order_id": transfer_id,
                    "quantity": quantity,
                    "tankers_required": tankers_required,
                    "available_tankers": AVAILABLE_TANKERS,
                    "sufficient":
                        tankers_required <= AVAILABLE_TANKERS,
                    "source_stp_name":
                        row.get("source_stp_name"),
                    "destination_stp_name":
                        row.get("destination_stp_name"),
                    "distance_km":
                        row.get("distance_km"),
                    "request_type":
                        "stp_transfer"
                }

                row["status"] = "Out for Delivery"

                row["tanker_status"] = "Out for Delivery"

            updated_rows.append(row)

    if not transfer_found:
        return f"Transfer {transfer_id} not found", 404

    if transfer_info is None:
        return "Transfer is not available for pickup", 400

    with open(
        STP_TRANSFERS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames
        )

        writer.writeheader()
        writer.writerows(updated_rows)

    return render_template(
        "tanker_summary.html",
        info=transfer_info,
        stp_id=None
    )

@app.route("/complete_transfer", methods=["POST"])
def complete_transfer():

    transfer_id = (
        request.form.get("transfer_id") or ""
    ).strip()

    if not transfer_id:
        return "No Transfer ID received", 400

    ensure_stp_transfers_file()

    stps = load_stps()

    updated_rows = []
    transfer_found = False
    completed = False

    with open(
        STP_TRANSFERS_FILE,
        "r",
        newline="",
        encoding="utf-8"
    ) as f:

        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames

        for row in reader:

            if (
                row.get("transfer_id", "").strip()
                != transfer_id
            ):
                updated_rows.append(row)
                continue

            transfer_found = True

            # Delivery can only happen after pickup
            if (
                row.get("status", "").strip()
                != "Out for Delivery"
            ):
                updated_rows.append(row)
                continue

            try:
                quantity_kld = float(
                    row.get("quantity_kld") or 0
                )
            except (TypeError, ValueError):
                return "Invalid transfer quantity", 400

            if quantity_kld <= 0:
                return "Transfer quantity must be greater than zero", 400

            quantity_mld = quantity_kld / 1000.0

            destination_stp = None

            for stp in stps:
                if (
                    str(stp.get("stp_id"))
                    == str(row.get("destination_stp_id"))
                ):
                    destination_stp = stp
                    break

            if destination_stp is None:
                return "Destination STP not found", 404

            # =================================================
            # ADD WATER TO DESTINATION STP
            # =================================================

            total_capacity = float(
                destination_stp.get(
                    "total_capacity_mld", 0
                ) or 0
            )

            available_capacity = float(
                destination_stp.get(
                    "available_capacity_mld", 0
                ) or 0
            )

            current_load = float(
                destination_stp.get(
                    "current_load_mld", 0
                ) or 0
            )

            destination_stp["available_capacity_mld"] = min(
                total_capacity,
                available_capacity + quantity_mld
            )

            destination_stp["current_load_mld"] = max(
                0.0,
                current_load - quantity_mld
            )

            # =================================================
            # COMPLETE TRANSFER
            # =================================================

            row["status"] = "Delivered"
            row["tanker_status"] = "Delivered"
            row["delivered_at"] = datetime.now().isoformat()

            completed = True

            updated_rows.append(row)

    if not transfer_found:
        return f"Transfer {transfer_id} not found", 404

    if not completed:
        return (
            "Transfer is not currently out for delivery",
            400
        )

    # Save destination STP capacity
    save_stps(stps)

    # Save transfer
    with open(
        STP_TRANSFERS_FILE,
        "w",
        newline="",
        encoding="utf-8"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=STP_TRANSFER_FIELDS
        )

        writer.writeheader()
        writer.writerows(updated_rows)

    return redirect(url_for("tanker_dashboard"))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(
        host="0.0.0.0",
        port=port,
        threaded=True
    )

from flask import Flask, jsonify, request, render_template, redirect, url_for
from flask import session
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

import threading
import uuid
from datetime import datetime, date, timedelta
import json
import os
import math
import requests
import pandas as pd
import csv
import osmnx as ox
import networkx as nx
from ml.predict_demand import predict_next_day, predict_week


def format_clean_address(address, lat, lon):
    try:
        place = (
            address.get("building") or
            address.get("amenity") or
            address.get("residential") or
            address.get("village") or
            address.get("hamlet")
        )
        area = address.get("suburb") or address.get("neighbourhood")
        road = address.get("road")
        district = address.get("state_district")
        state = address.get("state")
        pincode = address.get("postcode")
        formatted = ", ".join(filter(None, [place, area, road, district, state, pincode]))
        return formatted if formatted else f"{lat}, {lon}"
    except Exception as e:
        print("Format error:", e)
        return f"{lat}, {lon}"

app = Flask(__name__)

app.secret_key = "secret123"

# =========================================================
# LOAD ROAD NETWORK FOR A* ROUTING
# =========================================================
GRAPH_FILE = "bangalore_graph.graphml"

G = None

try:
    if os.path.exists(GRAPH_FILE):
        print("Loading saved road network...")
        G = ox.load_graphml(GRAPH_FILE)

        import random

        for u, v, k, data in G.edges(keys=True, data=True):
            traffic_factor = random.uniform(1.0, 3.0)

            data["traffic_factor"] = traffic_factor
            data["travel_cost"] = data["length"] * traffic_factor
    else:
        print("Skipping graph load (deployment)")
except Exception as e:
    print("Graph load skipped:", e)

print("Road network ready")

# =========================================================
# FILE PATHS
# =========================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

STP_FILE = os.path.join(BASE_DIR, "data", "stp_data.json")
STATUS_FILE = os.path.join(BASE_DIR, "data", "stp_status.json")
DATABASE_DIR = os.path.join(BASE_DIR, "database")
if not os.path.exists(DATABASE_DIR):
    os.makedirs(DATABASE_DIR)

# ✅ KEEP orders.csv INSIDE database/
ORDERS_FILE = os.path.join(DATABASE_DIR, "orders.csv")


# =========================================================

TANKER_REGISTRATIONS_FILE = os.path.join(
    DATABASE_DIR,
    "tanker_registrations.csv"
)

# =========================================================
# USER ACCOUNT DATABASE
# =========================================================

USERS_FILE = os.path.join(
    DATABASE_DIR,
    "users.xlsx"
)

users_lock = threading.Lock()
orders_lock = threading.Lock()

USER_FIELDS = [
    "user_id",
    "first_name",
    "last_name",
    "username",
    "mobile",
    "email",
    "password_hash",
    "role",
    "created_at",
    "account_status"
]

def ensure_users_file():
    """Create the Excel user database if it does not exist."""
    if not os.path.exists(USERS_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        sheet.append(USER_FIELDS)
        workbook.save(USERS_FILE)

def load_users():
    """Load all registered users from users.xlsx."""
    ensure_users_file()

    with users_lock:
        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]

        users = []
        for values in rows[1:]:
            user = {}
            for index, header in enumerate(headers):
                user[header] = values[index] if index < len(values) else ""
            users.append(user)

        return users

def append_user(user):
    """Append one user safely to users.xlsx."""
    ensure_users_file()

    with users_lock:
        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        # Ensure the expected header exists.
        existing_headers = [
            cell.value for cell in sheet[1]
        ]

        if existing_headers != USER_FIELDS:
            sheet.delete_rows(1, sheet.max_row)
            sheet.append(USER_FIELDS)

        sheet.append([
            user.get(field, "") for field in USER_FIELDS
        ])

        workbook.save(USERS_FILE)

ensure_users_file()

# =========================================================

# SYNTHETIC / DEMAND HEATMAP DATASET
# =========================================================

DEMAND_CSV_FILE = os.path.join(
    DATABASE_DIR,
    "synthetic_orders.csv"
)

# =========================================================

# ENSURE FILES EXIST
# =========================================================
if not os.path.exists(STATUS_FILE):
    with open(STATUS_FILE, "w") as f:
        json.dump({}, f)

if not os.path.exists(ORDERS_FILE):
    with open(ORDERS_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "order_id",
            "stp_id",
            "stp_name",
            "quantity_kld",
            "quality",
            "water_type",
            "distance_km",
            "location",
            "buyer_name",
            "buyer_phone",
            "status",
            "created_at"
        ])

ORDER_FIELDS = [
    "order_id",
    "stp_id",
    "stp_name",
    "quantity_kld",
    "quality",
    "water_type",
    "distance_km",
    "location",
    "buyer_user_id",
    "buyer_name",
    "buyer_phone",
    "status",
    "created_at",
    "payment_status",
    "accepted_at",
    "capacity_release_at",
    "capacity_released",
    "delivery_lat",
    "delivery_lon"
]

def ensure_orders_schema():
    """Add buyer_user_id to older orders.csv files without deleting existing orders."""
    if not os.path.exists(ORDERS_FILE) or os.path.getsize(ORDERS_FILE) == 0:
        with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
            writer.writeheader()
        return

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        existing_fields = reader.fieldnames or []
        rows = list(reader)

    if existing_fields == ORDER_FIELDS:
        return

    # Preserve every existing field and add the new account identifier.
    merged_fields = list(existing_fields)
    if "buyer_user_id" not in merged_fields:
        insert_at = merged_fields.index("buyer_name") if "buyer_name" in merged_fields else len(merged_fields)
        merged_fields.insert(insert_at, "buyer_user_id")

    # Keep the new canonical order.
    for field in ORDER_FIELDS:
        if field not in merged_fields:
            merged_fields.append(field)

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()

        for row in rows:
            normalized = {field: row.get(field, "") for field in ORDER_FIELDS}
            writer.writerow(normalized)

ensure_orders_schema()


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def load_stps():
    with open(STP_FILE) as f:
        data = json.load(f)
        return data.get("stps", [])

def save_stps(stps):
    with open(STP_FILE, "w") as f:
        json.dump({"stps": stps}, f, indent=4)

def auto_reset_capacity():
    """Release STP capacity for accepted orders exactly 24 hours after acceptance."""
    now = datetime.now()
    stps = load_stps()
    stps_changed = False
    orders_changed = False

    if not os.path.exists(ORDERS_FILE):
        return

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        orders = list(reader)

    for row in orders:
        status = (row.get("status") or "").strip()
        if status not in {"Accepted", "Out for Delivery", "Delivered"}:
            continue

        release_at_raw = (row.get("capacity_release_at") or "").strip()
        if not release_at_raw:
            accepted_at_raw = (row.get("accepted_at") or "").strip() or (row.get("created_at") or "").strip()
            try:
                accepted_at = datetime.fromisoformat(accepted_at_raw)
                release_at = accepted_at + timedelta(hours=24)
                row["accepted_at"] = accepted_at.isoformat()
                row["capacity_release_at"] = release_at.isoformat()
                row["capacity_released"] = row.get("capacity_released") or "False"
                release_at_raw = release_at.isoformat()
                orders_changed = True
            except (TypeError, ValueError):
                continue

        if str(row.get("capacity_released", "")).strip().lower() == "true":
            continue

        try:
            release_at = datetime.fromisoformat(release_at_raw)
        except (TypeError, ValueError):
            continue

        if now < release_at:
            continue

        quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
        for stp in stps:
            if str(stp.get("stp_id")) == str(row.get("stp_id")):
                total_capacity = float(stp.get("total_capacity_mld") or 0)
                available_capacity = float(stp.get("available_capacity_mld", total_capacity) or 0)
                stp["available_capacity_mld"] = min(total_capacity, available_capacity + quantity_mld)
                stp["current_load_mld"] = max(0.0, float(stp.get("current_load_mld", 0) or 0) - quantity_mld)
                stps_changed = True
                row["capacity_released"] = "True"
                orders_changed = True
                break

    if stps_changed:
        save_stps(stps)
    if orders_changed:
        with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
            writer.writeheader()
            for row in orders:
                writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})


def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) *
         math.cos(math.radians(lat2)) *
         math.sin(dlon/2)**2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

# =========================================================
# A* DISTANCE FUNCTION
# =========================================================
def astar_distance(lat1, lon1, lat2, lon2):
    
    if G is None:
        print("Using fallback distance")
        return haversine(lat1, lon1, lat2, lon2)

    try:
        start_node = ox.distance.nearest_nodes(G, lon1, lat1)
        end_node = ox.distance.nearest_nodes(G, lon2, lat2)

        distance_meters = nx.astar_path_length(G, start_node, end_node, weight="travel_cost")
        return round(distance_meters / 1000, 2)

    except Exception as e:
        print("A* failed, fallback:", e)
        return haversine(lat1, lon1, lat2, lon2)
    
    from itertools import islice

    def get_alternative_routes(lat1, lon1, lat2, lon2):

        start_node = ox.distance.nearest_nodes(G, lon1, lat1)
        end_node = ox.distance.nearest_nodes(G, lon2, lat2)

        routes = list(
            islice(
                nx.shortest_simple_paths(
                    G,
                    start_node,
                    end_node,
                    weight="travel_cost"
                ),
                3
            )
        )

        return routes

    print("Running A* routing...")

    start_node = ox.distance.nearest_nodes(G, lon1, lat1)
    end_node = ox.distance.nearest_nodes(G, lon2, lat2)

    distance_meters = nx.astar_path_length(G, start_node, end_node, weight="length")

    distance_km = distance_meters / 1000

    print(f"A* distance: {distance_km:.2f} km")

    return distance_km

# =========================================================
# HOME + LOGIN
# =========================================================

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        login_identifier = request.form.get("login_identifier", "").strip()
        password = request.form.get("password", "")

        if not login_identifier or not password:
            return render_template(
                "login.html",
                login_error="Please enter your username/email and password."
            )

        users = load_users()
        matched_user = None

        for user in users:
            username = str(user.get("username") or "").strip().lower()
            email = str(user.get("email") or "").strip().lower()

            if login_identifier.lower() in [username, email]:
                matched_user = user
                break

        if matched_user is None:
            return render_template(
                "login.html",
                login_error="Invalid username/email or password."
            )

        if str(matched_user.get("account_status") or "").strip().lower() != "active":
            return render_template(
                "login.html",
                login_error="Your account is not active. Please contact the administrator."
            )

        password_hash = str(matched_user.get("password_hash") or "")

        if not password_hash or not check_password_hash(password_hash, password):
            return render_template(
                "login.html",
                login_error="Invalid username/email or password."
            )

        # Each browser receives its own independent Flask session.
        session.clear()

        session["user_id"] = str(matched_user.get("user_id") or "")
        session["first_name"] = str(matched_user.get("first_name") or "")
        session["last_name"] = str(matched_user.get("last_name") or "")
        session["username"] = str(matched_user.get("username") or "")

        session["user_name"] = (
            f"{matched_user.get('first_name', '')} "
            f"{matched_user.get('last_name', '')}"
        ).strip()

        session["user_phone"] = str(matched_user.get("mobile") or "")
        session["user_email"] = str(matched_user.get("email") or "")
        session["role"] = str(matched_user.get("role") or "").strip().lower()

        # Keep the existing buyer session variables.
        if session["role"] == "demand":
            session["buyer_name"] = session["user_name"]
            session["buyer_phone"] = session["user_phone"]
            return redirect(url_for("demand"))

        if session["role"] == "stp":
            return redirect(url_for("supply"))

        if session["role"] == "tanker":
            # Keep the existing tanker dashboard flow.
            # If this account is linked to an approved tanker operator,
            # its operator id can be used later for tanker-specific assignments.
            session["tanker_operator_id"] = session["user_id"]
            session["tanker_operator_name"] = session["user_name"]
            return redirect(url_for("tanker_dashboard"))

        if session["role"] == "admin":
            return redirect(url_for("admin_dashboard"))

        session.clear()

        return render_template(
            "login.html",
            login_error="Your account has an invalid role."
        )

    return render_template("login.html")


@app.route('/signup', methods=['GET', 'POST'])
def signup():

    if request.method == 'POST':

        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        mobile = request.form.get("mobile", "").strip()
        email = request.form.get("email", "").strip().lower()
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        role = request.form.get("role", "").strip().lower()

        allowed_roles = {"demand", "stp", "tanker"}

        if not all([
            first_name,
            last_name,
            mobile,
            email,
            username,
            password,
            confirm_password,
            role
        ]):
            return render_template(
                "signup.html",
                signup_error="Please fill in all fields."
            )

        if role not in allowed_roles:
            return render_template(
                "signup.html",
                signup_error="Please select a valid account type."
            )

        if password != confirm_password:
            return render_template(
                "signup.html",
                signup_error="Passwords do not match."
            )

        if len(password) < 8:
            return render_template(
                "signup.html",
                signup_error="Password must be at least 8 characters long."
            )

        users = load_users()

        for user in users:
            existing_username = str(user.get("username") or "").strip().lower()
            existing_email = str(user.get("email") or "").strip().lower()

            if username == existing_username:
                return render_template(
                    "signup.html",
                    signup_error="That username is already registered."
                )

            if email == existing_email:
                return render_template(
                    "signup.html",
                    signup_error="That email address is already registered."
                )

            if mobile == str(user.get("mobile") or "").strip():
                return render_template(
                    "signup.html",
                    signup_error="That mobile number is already registered."
                )

        user_id = "USR-" + uuid.uuid4().hex[:10].upper()

        new_user = {
            "user_id": user_id,
            "first_name": first_name,
            "last_name": last_name,
            "username": username,
            "mobile": mobile,
            "email": email,
            "password_hash": generate_password_hash(password),
            "role": role,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "account_status": "active"
        }

        append_user(new_user)

        return redirect(
            url_for(
                "login",
                signup_success="Account created successfully. Please log in."
            )
        )

    return render_template("signup.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/delete_account", methods=["POST"])
def delete_account():

    # User must be logged in
    if not session.get("user_id"):
        return redirect(url_for("login"))

    user_id = str(session.get("user_id") or "")

    if not user_id:
        return redirect(url_for("login"))

    # Make sure users.xlsx exists
    ensure_users_file()

    with users_lock:

        workbook = load_workbook(USERS_FILE)
        sheet = workbook["Users"]

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

        # Find user_id column
        if "user_id" not in headers:
            workbook.close()
            return "User ID column not found in users.xlsx", 500

        user_id_column = headers.index("user_id") + 1

        user_found = False

        # Find and delete the logged-in user's row
        for row in range(2, sheet.max_row + 1):

            current_user_id = str(
                sheet.cell(
                    row=row,
                    column=user_id_column
                ).value or ""
            ).strip()

            if current_user_id == user_id:

                sheet.delete_rows(row, 1)
                user_found = True
                break

        workbook.save(USERS_FILE)
        workbook.close()

    # Clear the current login session
    session.clear()

    if user_found:
        return redirect(
            url_for(
                "login",
                account_deleted="Account deleted successfully."
            )
        )

    return redirect(url_for("login"))


# =========================================================
# CURRENT LOGGED-IN USER
# =========================================================

@app.route("/profile")
def profile():

    # Check whether a user is logged in
    if not session.get("user_id"):
        return redirect(url_for("login"))

    # Get the complete user record from users.xlsx
    users = load_users()
    logged_in_user = None

    for user in users:
        if str(user.get("user_id") or "") == str(session.get("user_id") or ""):
            logged_in_user = user
            break

    if logged_in_user is None:
        session.clear()
        return redirect(url_for("login"))

    return render_template(
        "profile.html",
        user={
            "user_id": logged_in_user.get("user_id", ""),
            "first_name": logged_in_user.get("first_name", ""),
            "last_name": logged_in_user.get("last_name", ""),
            "name": f"{logged_in_user.get('first_name', '')} {logged_in_user.get('last_name', '')}".strip(),
            "username": logged_in_user.get("username", ""),
            "mobile": logged_in_user.get("mobile", ""),
            "email": logged_in_user.get("email", ""),
            "role": logged_in_user.get("role", ""),
            "created_at": logged_in_user.get("created_at", ""),
            "account_status": logged_in_user.get("account_status", "")
        }
    )
    
@app.route("/api/current_user")
def current_user():
    """Return only the user stored in this browser's Flask session."""
    user_id = session.get("user_id")

    if not user_id:
        return jsonify({
            "logged_in": False,
            "initials": "👤",
            "name": "Guest",
            "role": ""
        })

    first_name = str(session.get("first_name") or "").strip()
    last_name = str(session.get("last_name") or "").strip()

    initials = ""
    if first_name:
        initials += first_name[0].upper()
    if last_name:
        initials += last_name[0].upper()
    if not initials:
        initials = "👤"

    return jsonify({
        "logged_in": True,
        "first_name": first_name,
        "last_name": last_name,
        "name": str(session.get("user_name") or "").strip(),
        "role": str(session.get("role") or "").strip(),
        "initials": initials
    })


@app.route("/tanker/register")
def tanker_register():
    return render_template("tanker_register.html")

@app.route("/tanker/status", methods=["GET", "POST"])
def tanker_status():

    if request.method == "POST":

        operator_id = request.form.get("operator_id", "").strip()
        phone = request.form.get("phone", "").strip()

        operator = None

        if os.path.exists(TANKER_REGISTRATIONS_FILE):

            with open(
                TANKER_REGISTRATIONS_FILE,
                "r",
                newline="",
                encoding="utf-8"
            ) as f:

                reader = csv.DictReader(f)

                for row in reader:

                    if (
                        row.get("operator_id", "").strip() == operator_id
                        and
                        row.get("phone", "").strip() == phone
                    ):
                        operator = row
                        break

        return render_template(
            "tanker_status.html",
            operator=operator,
            searched=True
        )

    return render_template(
        "tanker_status.html",
        operator=None,
        searched=False
    )

@app.route("/tanker/register/contracted", methods=["GET", "POST"])
def tanker_register_contracted():

    if request.method == "POST":

        # Operator details
        operator_name = request.form.get("operator_name")
        phone = request.form.get("phone")
        email = request.form.get("email")

        # Contract details
        contract_id = request.form.get("contract_id")
        contract_start = request.form.get("contract_start")
        contract_end = request.form.get("contract_end")

        # Tanker details
        registration_no = request.form.get("registration_no")
        capacity = request.form.get("capacity")
        vehicle_model = request.form.get("vehicle_model")

        # CSV file location
        csv_file = os.path.join(
            app.root_path,
            "database",
            "tanker_registrations.csv"
        )

        # Generate operator ID
        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
            existing_df = pd.read_csv(csv_file)
            operator_number = len(existing_df) + 1
        else:
            operator_number = 1

        operator_id = f"OP-BLR-{operator_number:04d}"

        # Create registration record
        new_operator = {
            "operator_id": operator_id,
            "operator_name": operator_name,
            "operator_type": "contracted",
            "phone": phone,
            "email": email,
            "area": "",
            "pincode": "",
            "contract_id": contract_id,
            "contract_start": contract_start,
            "contract_end": contract_end,
            "tanker_registration_no": registration_no,
            "tanker_capacity_kl": capacity,
            "vehicle_model": vehicle_model,
            "water_type_supported": "",
            "service_radius_km": "",
            "verification_status": "pending",
            "registration_date": date.today().isoformat()
        }

        # Convert to DataFrame
        new_df = pd.DataFrame([new_operator])

        # Save to CSV
        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
            new_df.to_csv(
                csv_file,
                mode="a",
                header=False,
                index=False
            )
        else:
            new_df.to_csv(
                csv_file,
                mode="w",
                header=True,
                index=False
            )

        print("NEW CONTRACTED TANKER OPERATOR REGISTERED")
        print(new_operator)
        print("DATA SAVED TO:", csv_file)

        return render_template(
            "registration_success.html",
        operator_id=operator_id,
        operator_type="Existing Purvankara Partner"
)

    return render_template("tanker_register_contracted.html")

@app.route("/tanker/register/independent", methods=["GET", "POST"])
def tanker_register_independent():

    if request.method == "POST":

        # Get the submitted form data
        operator_name = request.form.get("operator_name")
        phone = request.form.get("phone")
        email = request.form.get("email")

        area = request.form.get("area")
        pincode = request.form.get("pincode")

        registration_no = request.form.get("registration_no")
        capacity = request.form.get("capacity")
        vehicle_model = request.form.get("vehicle_model")

        water_type = request.form.get("water_type")
        radius = request.form.get("radius")

        # CSV file location
        csv_file = os.path.join(
            app.root_path,
            "database",
            "tanker_registrations.csv"
        )
        print("CSV PATH:", csv_file)
        print("CSV EXISTS:", os.path.exists(csv_file))

        # Generate a new operator ID
        if os.path.exists(csv_file):

            existing_df = pd.read_csv(csv_file)

            operator_number = len(existing_df) + 1

        else:
            operator_number = 1

        operator_id = f"OP-BLR-{operator_number:04d}"

        # Create the new registration
        new_operator = {
            "operator_id": operator_id,
            "operator_name": operator_name,
            "operator_type": "independent",
            "phone": phone,
            "email": email,
            "area": area,
            "pincode": pincode,
            "contract_id": "",
            "contract_start": "",
            "contract_end": "",
            "tanker_registration_no": registration_no,
            "tanker_capacity_kl": capacity,
            "vehicle_model": vehicle_model,
            "water_type_supported": water_type,
            "service_radius_km": radius,
            "verification_status": "pending",
            "registration_date": date.today().isoformat()
        }

        # Add the registration to the CSV
        new_df = pd.DataFrame([new_operator])

        if os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
                new_df.to_csv(
                    csv_file,
                    mode="a",
                    header=False,
                    index=False
                )
        else:
                new_df.to_csv(
                    csv_file,
                    mode="w",
                    header=True,
                    index=False
                )

        print("DATA SAVED TO:", csv_file)

        print("NEW TANKER OPERATOR REGISTERED")
        print(new_operator)

        return render_template(
            "registration_success.html",
        operator_id=operator_id,
        operator_type="Independent Operator"
)

    return render_template("tanker_register_independent.html")



# =========================================================
# ADMIN DASHBOARD
# =========================================================

@app.route("/admin")
def admin_dashboard():
    stps = load_stps()

    orders = []
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)
            orders = list(reader)


    return render_template("admin.html", stps=stps, orders=orders)


    tanker_operators = []

    if os.path.exists(TANKER_REGISTRATIONS_FILE):
        with open(TANKER_REGISTRATIONS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            tanker_operators = list(reader)

    total_tanker_operators = len(tanker_operators)

    pending_tanker_operators = sum(
        1 for operator in tanker_operators
        if operator.get("verification_status", "").strip().lower() == "pending"
    )

    approved_tanker_operators = sum(
        1 for operator in tanker_operators
        if operator.get("verification_status", "").strip().lower() == "approved"
    )

    rejected_tanker_operators = sum(
        1 for operator in tanker_operators
        if operator.get("verification_status", "").strip().lower() == "rejected"
    )

    return render_template(
        "admin.html",
        stps=stps,
        orders=orders,
        tanker_operators=tanker_operators,
        total_tanker_operators=total_tanker_operators,
        pending_tanker_operators=pending_tanker_operators,
        approved_tanker_operators=approved_tanker_operators,
        rejected_tanker_operators=rejected_tanker_operators
    )

@app.route("/admin/tanker/<operator_id>/status/<status>")
def update_tanker_status(operator_id, status):

    # Only allow valid statuses
    if status not in ["approved", "rejected"]:
        return redirect("/admin")

    rows = []

    if os.path.exists(TANKER_REGISTRATIONS_FILE):

        with open(
            TANKER_REGISTRATIONS_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as f:

            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            rows = list(reader)

        # Update the matching operator
        for operator in rows:

            if operator.get("operator_id") == operator_id:
                operator["verification_status"] = status
                break

        # Save updated CSV
        with open(
            TANKER_REGISTRATIONS_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as f:

            writer = csv.DictWriter(
                f,
                fieldnames=fieldnames
            )

            writer.writeheader()
            writer.writerows(rows)

    return redirect("/admin")


# =========================
# ADD STP
# =========================
@app.route("/admin/add_stp", methods=["POST"])
def add_stp():
    stps = load_stps()

    new_stp = {
        "stp_id": request.form["id"],
        "stp_name": request.form["name"],
        "latitude": float(request.form["lat"]),
        "longitude": float(request.form["lon"]),
        "technology": "Manual",
        "total_capacity_mld": float(request.form["capacity"]),
        "current_load_mld": 0.0,
        "available_capacity_mld": float(request.form["capacity"]),
        "treatment_cost_per_kl": 5.0,
        "quality_grade": "General",

        "last_reset_date": date.today().isoformat(),
        "last_reset_at": datetime.now().isoformat()
    }

    stps.append(new_stp)
    save_stps(stps)

    return redirect(url_for("admin_dashboard"))


# =========================
# DELETE STP
# =========================
@app.route("/admin/delete_stp/<stp_id>")
def delete_stp(stp_id):
    stps = load_stps()

    stps = [s for s in stps if str(s["stp_id"]) != str(stp_id)]

    save_stps(stps)

    return redirect(url_for("admin_dashboard"))

# =========================================================
# DEMAND SIDE
# =========================================================

@app.route('/demand')
def demand():
    payment_success = request.args.get("payment_success")

    return render_template(
        "demand.html",
        payment_success=payment_success
    )

@app.route("/track")
def track_page():
    return render_template("track.html")

@app.route("/api/stps")
def api_stps():

    return jsonify(load_stps())


    auto_reset_capacity()
    return jsonify(load_stps())


# =========================================================
# DEMAND HEATMAP API
# =========================================================

@app.route("/api/demand_heatmap")
def demand_heatmap():

    demand_data = []

    if not os.path.exists(DEMAND_CSV_FILE):
        return jsonify({
            "error": "synthetic_orders.csv not found"
        }), 404

    try:

        with open(
            DEMAND_CSV_FILE,
            "r",
            newline="",
            encoding="utf-8"
        ) as file:

            reader = csv.DictReader(file)

            for row in reader:

                try:

                    latitude = float(row["latitude"])
                    longitude = float(row["longitude"])
                    quantity = float(row["quantity_kld"])

                    demand_data.append({
                        "latitude": latitude,
                        "longitude": longitude,
                        "quantity_kld": quantity
                    })

                except (
                    KeyError,
                    ValueError,
                    TypeError
                ):
                    # Ignore malformed rows
                    continue

        return jsonify(demand_data)

    except Exception as e:

        print("Demand heatmap error:", e)

        return jsonify({
            "error": "Could not read synthetic_orders.csv"
        }), 500
    

@app.route("/api/search_place")
def api_search_place():

    auto_reset_capacity()

    place = request.args.get("place")
    lat = request.args.get("lat")
    lon = request.args.get("lon")

    # Keep the existing location behavior: typed location or live location.
    if lat and lon:
        lat = float(lat)
        lon = float(lon)

        reverse_url = (
            f"https://nominatim.openstreetmap.org/reverse"
            f"?format=json&lat={lat}&lon={lon}"
        )
        try:
            response = requests.get(
                reverse_url,
                headers={"User-Agent": "wastewater-app"},
                timeout=5
            )
            reverse_data = response.json()
        except Exception as e:
            print("Reverse API failed:", e)
            reverse_data = {}

        address = reverse_data.get("address", {})
        location_name = format_clean_address(address, lat, lon)

        if not location_name or not location_name.strip():
            location_name = reverse_data.get(
                "display_name",
                f"{lat}, {lon}"
            )

        print("Using LIVE coordinates:", lat, lon)

    elif place and place != "Using Live Location":
        search_queries = [
            f"{place}, Bengaluru, Karnataka, India",
            f"{place}, Bangalore, Karnataka, India",
            f"{place}, Karnataka, India",
        ]

        geo_data = []

        for search_place in search_queries:
            geo_url = (
                "https://nominatim.openstreetmap.org/search"
                f"?format=json&limit=1&q="
                f"{requests.utils.quote(search_place)}"
            )

            try:
                response = requests.get(
                    geo_url,
                    headers={"User-Agent": "wastewater-app"},
                    timeout=8
                )

                if response.ok:
                    geo_data = response.json()

                if geo_data:
                    break

            except Exception as e:
                print("Location search failed:", e)

        if not geo_data:
            return jsonify({
                "error": f"Unable to find location: {place}"
            }), 404

        lat = float(geo_data[0]["lat"])
        lon = float(geo_data[0]["lon"])
        location_name = place

    else:
        return jsonify({"error": "No location provided"}), 400

    try:
        required_kld = float(request.args.get("required_kld", 0) or 0)
    except (TypeError, ValueError):
        required_kld = 0.0

    required_quality = str(
        request.args.get("quality") or ""
    ).strip()

    required_type = str(
        request.args.get("type") or ""
    ).strip()

    required_mld = required_kld / 1000.0

    # Remember the exact location used for this Demand search. The existing
    # booking page may send only the displayed address/name when the user
    # clicks Book Order, so create_order can still use the exact coordinates.
    session["last_demand_location"] = {
        "latitude": lat,
        "longitude": lon,
        "name": location_name
    }

    stps = load_stps()
    nearby = []

    requested_quality = required_quality.lower()
    requested_type = required_type.lower()

    for stp in stps:

        # STP must have coordinates.
        try:
            stp_lat = float(stp.get("latitude"))
            stp_lon = float(stp.get("longitude"))
        except (TypeError, ValueError):
            continue

        # ---------------------------------------------------------
        # 1. CAPACITY MATCH
        # ---------------------------------------------------------
        try:
            raw_available = stp.get("available_capacity_mld")

            if raw_available not in (None, ""):
                available_capacity = float(raw_available)
            else:
                total_capacity = float(
                    stp.get("total_capacity_mld", 0) or 0
                )
                current_load = float(
                    stp.get("current_load_mld", 0) or 0
                )
                available_capacity = max(
                    0.0,
                    total_capacity - current_load
                )
        except (TypeError, ValueError):
            available_capacity = 0.0

        if required_mld > 0 and available_capacity < required_mld:
            continue

        # ---------------------------------------------------------
        # 2. QUALITY MATCH
        # ---------------------------------------------------------
        stp_quality = str(
            stp.get("quality_grade") or ""
        ).strip().lower()

        # If an STP has a quality value, it must match the user's
        # requested quality. Empty STP quality remains compatible,
        # matching the existing STP acceptance logic.
        if (
            requested_quality
            and stp_quality
            and requested_quality != stp_quality
        ):
            continue

        # ---------------------------------------------------------
        # 3. WATER TYPE MATCH
        # ---------------------------------------------------------
        stp_type = str(
            stp.get("water_type") or ""
        ).strip().lower()

        # Some existing STP records do not contain water_type.
        # Do NOT reject those records just because the Demand page
        # selected "Treated". They are treated STPs in this system,
        # and the existing acceptance logic treats an empty type
        # as compatible.
        if (
            requested_type
            and stp_type
            and requested_type != stp_type
        ):
            continue

        # ---------------------------------------------------------
        # 4. LOCATION MATCH
        # ---------------------------------------------------------
        straight_distance = haversine(
            lat,
            lon,
            stp_lat,
            stp_lon
        )

        # Keep the STP search within a practical Bengaluru range.
        if straight_distance > 100:
            continue

        # Use the existing A* route distance where available.
        try:
            route_distance = float(
                astar_distance(
                    lat,
                    lon,
                    stp_lat,
                    stp_lon
                )
            )
        except Exception as e:
            print(
                f"A* distance failed for {stp.get('stp_id')}:",
                e
            )
            route_distance = None

        # If A* cannot calculate a route, don't hide a valid STP.
        # The Demand page itself uses OSRM to draw the actual road route.
        if (
            route_distance is None
            or route_distance <= 0
            or route_distance > 100
        ):
            route_distance = straight_distance

        if route_distance > 100:
            continue

        stp_copy = stp.copy()
        stp_copy["latitude"] = stp_lat
        stp_copy["longitude"] = stp_lon
        stp_copy["distance_km"] = round(
            route_distance,
            2
        )
        stp_copy["available_capacity_mld"] = round(
            available_capacity,
            6
        )

        stp_copy["match_reason"] = (
            "Demand matched: capacity + quality + "
            "water type + location"
        )

        nearby.append(stp_copy)

    # IMPORTANT:
    # Select the nearest STP ONLY from STPs that satisfy the demand.
    nearby.sort(
        key=lambda x: float(x.get("distance_km", 999999))
    )

    nearest = nearby[0] if nearby else None

    if not nearest:
        return jsonify({
            "searched_location": {
                "name": location_name,
                "latitude": lat,
                "longitude": lon
            },
            "nearest_stp": None,
            "all_stps": [],
            "matching_error": (
                "No STP currently satisfies the requested "
                "quantity, quality, water type and location."
            )
        })

    print(
        "MATCHED STP:",
        nearest.get("stp_id"),
        nearest.get("stp_name"),
        "| Demand:",
        required_kld,
        "KLD",
        required_quality,
        required_type,
        "| Distance:",
        nearest.get("distance_km"),
        "km"
    )

    return jsonify({
        "searched_location": {
            "name": location_name,
            "latitude": lat,
            "longitude": lon
        },
        "nearest_stp": nearest,
        "all_stps": nearby
    })


def resolve_delivery_coordinates(data):
    """
    Get the exact delivery coordinates supplied by the Demand page.
    Supports several common key names so existing frontend code does not
    need to be rewritten. For older bookings that only send an address,
    use Nominatim once at order creation and persist the result.
    """
    lat_keys = ("delivery_lat", "latitude", "lat", "buyer_lat")
    lon_keys = ("delivery_lon", "longitude", "lon", "lng", "buyer_lon")

    lat = next((data.get(k) for k in lat_keys if data.get(k) not in (None, "")), None)
    lon = next((data.get(k) for k in lon_keys if data.get(k) not in (None, "")), None)

    try:
        if lat is not None and lon is not None:
            return float(lat), float(lon)
    except (TypeError, ValueError):
        pass

    # Fallback only when the Demand page did not send coordinates.
    location = str(data.get("location") or "").strip()
    if not location:
        return None, None

    try:
        geo_url = (
            "https://nominatim.openstreetmap.org/search"
            f"?format=json&limit=1&q={requests.utils.quote(location)}"
        )
        response = requests.get(
            geo_url,
            headers={"User-Agent": "wastewater-app"},
            timeout=8
        )
        geo_data = response.json()
        if geo_data:
            return float(geo_data[0]["lat"]), float(geo_data[0]["lon"])
    except Exception as e:
        print("Delivery location geocoding failed:", e)

    return None, None


@app.route("/create_order", methods=["POST"])
def create_order():
    data = request.json or {}

    required = [
        "stp_id", "stp_name", "quantity_kld", "quality",
        "water_type", "distance_km", "location"
    ]
    missing = [key for key in required if key not in data]
    if missing:
        return jsonify({"error": "Missing fields", "fields": missing}), 400

    # -------------------------------------------------------------
    # EXACT DEMAND LOCATION
    # -------------------------------------------------------------
    # First use coordinates sent by the frontend.
    delivery_lat, delivery_lon = resolve_delivery_coordinates(data)

    # If the existing Demand page only sends the displayed location,
    # reuse the exact coordinates from the user's most recent search/live
    # location instead of geocoding an approximate address.
    if delivery_lat is None or delivery_lon is None:
        last_location = session.get("last_demand_location") or {}

        try:
            if (
                last_location.get("latitude") is not None
                and last_location.get("longitude") is not None
            ):
                delivery_lat = float(last_location["latitude"])
                delivery_lon = float(last_location["longitude"])
        except (TypeError, ValueError):
            delivery_lat = delivery_lon = None

    if delivery_lat is None or delivery_lon is None:
        return jsonify({
            "error": "Delivery location could not be resolved."
        }), 422

    # -------------------------------------------------------------
    # DEMAND REQUIREMENTS
    # -------------------------------------------------------------
    try:
        requested_kld = float(data.get("quantity_kld") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid quantity"}), 400

    if requested_kld <= 0:
        return jsonify({"error": "Quantity must be greater than zero"}), 400

    requested_mld = requested_kld / 1000.0
    requested_quality = str(
        data.get("quality") or ""
    ).strip().lower()
    requested_type = str(
        data.get("water_type") or ""
    ).strip().lower()

    # -------------------------------------------------------------
    # FINAL SERVER-SIDE STP MATCH
    # -------------------------------------------------------------
    # Do not blindly trust the STP id returned by the browser.
    # Recalculate the best feasible STP using the same rules as
    # /api/search_place.
    feasible_stps = []

    for stp in load_stps():

        try:
            stp_lat = float(stp.get("latitude"))
            stp_lon = float(stp.get("longitude"))
        except (TypeError, ValueError):
            continue

        # Capacity: use explicit available capacity when present;
        # otherwise calculate total - current load.
        try:
            raw_available = stp.get("available_capacity_mld")

            if raw_available not in (None, ""):
                available_capacity = float(raw_available)
            else:
                total_capacity = float(
                    stp.get("total_capacity_mld", 0) or 0
                )
                current_load = float(
                    stp.get("current_load_mld", 0) or 0
                )
                available_capacity = max(
                    0.0,
                    total_capacity - current_load
                )
        except (TypeError, ValueError):
            available_capacity = 0.0

        if requested_mld > 0 and available_capacity < requested_mld:
            continue

        # Quality: match when the STP record contains a quality value.
        stp_quality = str(
            stp.get("quality_grade") or ""
        ).strip().lower()

        if (
            requested_quality
            and stp_quality
            and requested_quality != stp_quality
        ):
            continue

        # Water type: match when the STP record contains a type.
        # Empty type remains compatible with existing STP records.
        stp_type = str(
            stp.get("water_type") or ""
        ).strip().lower()

        if (
            requested_type
            and stp_type
            and requested_type != stp_type
        ):
            continue

        # Location feasibility.
        straight_distance = haversine(
            delivery_lat,
            delivery_lon,
            stp_lat,
            stp_lon
        )

        if straight_distance > 100:
            continue

        # The deployment environment may not have the A* graph.
        # Fall back to haversine instead of rejecting a valid STP.
        try:
            route_distance = float(
                astar_distance(
                    delivery_lat,
                    delivery_lon,
                    stp_lat,
                    stp_lon
                )
            )
        except Exception as e:
            print(
                f"A* distance failed for {stp.get('stp_id')}: {e}"
            )
            route_distance = None

        if (
            route_distance is None
            or route_distance <= 0
            or route_distance > 100
        ):
            route_distance = straight_distance

        if route_distance > 100:
            continue

        candidate = stp.copy()
        candidate["latitude"] = stp_lat
        candidate["longitude"] = stp_lon
        candidate["distance_km"] = round(
            route_distance,
            2
        )
        candidate["available_capacity_mld"] = round(
            available_capacity,
            6
        )

        feasible_stps.append(candidate)

    if not feasible_stps:
        print(
            "ORDER BLOCKED: no feasible STP for",
            requested_kld,
            "KLD",
            requested_quality,
            requested_type
        )

        return jsonify({
            "error": (
                "No STP satisfies your requested quantity, water quality, "
                "water type and delivery location."
            )
        }), 422

    # Closest STP among ONLY the STPs that satisfy the demand.
    matching_stp = min(
        feasible_stps,
        key=lambda stp: float(stp["distance_km"])
    )

    order_id = "ORD-" + uuid.uuid4().hex[:10].upper()

    # Use the server-selected STP, not a random/browser-selected STP.
    data["stp_id"] = matching_stp["stp_id"]
    data["stp_name"] = matching_stp["stp_name"]
    data["distance_km"] = matching_stp["distance_km"]

    row = {
        "order_id": order_id,
        "stp_id": data["stp_id"],
        "stp_name": data["stp_name"],
        "quantity_kld": data["quantity_kld"],
        "quality": data["quality"],
        "water_type": data["water_type"],
        "distance_km": data["distance_km"],
        "location": data["location"],
        "buyer_user_id": session.get("user_id") or "",
        "buyer_name": (
            session.get("buyer_name")
            or session.get("user_name")
            or "Unknown"
        ),
        "buyer_phone": (
            session.get("buyer_phone")
            or session.get("user_phone")
            or "N/A"
        ),
        "status": "Pending",
        "created_at": datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        "payment_status": "Pending",
        "accepted_at": "",
        "capacity_release_at": "",
        "capacity_released": "False",
        "delivery_lat": delivery_lat,
        "delivery_lon": delivery_lon
    }

    with open(ORDERS_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=ORDER_FIELDS
        )
        writer.writerow(row)

    print(
        "ORDER CREATED:",
        order_id,
        "| STP:",
        matching_stp.get("stp_id"),
        matching_stp.get("stp_name"),
        "| Delivery:",
        delivery_lat,
        delivery_lon
    )

    return jsonify({
        "message": "Order created successfully",
        "order_id": order_id,
        "stp_id": matching_stp["stp_id"],
        "stp_name": matching_stp["stp_name"]
    })


@app.route("/api/order_tracking/<order_id>")
def api_order_tracking(order_id):
    if not session.get("user_id"):
        return jsonify({"success": False, "error": "Login required"}), 401

    order = None
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if str(row.get("order_id", "")).strip() == str(order_id).strip():
                    order = row
                    break

    if not order:
        return jsonify({"success": False, "error": "Order not found"}), 404

    current_user_id = str(session.get("user_id") or "").strip()
    buyer_user_id = str(order.get("buyer_user_id") or "").strip()
    user_role = str(session.get("role") or "").strip().lower()

    if user_role == "demand":
        if current_user_id and buyer_user_id and current_user_id != buyer_user_id:
            return jsonify({"success": False, "error": "Unauthorized"}), 403

    elif user_role == "stp":
        selected_stp_id = str(
            request.args.get("stp_id")
            or session.get("selected_stp_id")
            or ""
        ).strip()

        if (
            selected_stp_id
            and str(order.get("stp_id") or "").strip() != selected_stp_id
        ):
            return jsonify({"success": False, "error": "Unauthorized"}), 403

    else:
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    stp_lat = stp_lon = None
    for stp in load_stps():
        if str(stp.get("stp_id", "")).strip() == str(order.get("stp_id", "")).strip():
            stp_lat = stp.get("latitude")
            stp_lon = stp.get("longitude")
            break

    try:
        delivery_lat = float(order.get("delivery_lat"))
        delivery_lon = float(order.get("delivery_lon"))
    except (TypeError, ValueError):
        delivery_lat, delivery_lon = resolve_delivery_coordinates(order)

    if stp_lat is None or stp_lon is None:
        return jsonify({"success": False, "error": "STP coordinates unavailable"}), 404

    if delivery_lat is None or delivery_lon is None:
        return jsonify({
            "success": False,
            "error": "Exact delivery location is unavailable for this order"
        }), 422

    return jsonify({
        "success": True,
        "order_id": order.get("order_id"),
        "status": order.get("status"),
        "stp": {
            "id": order.get("stp_id"),
            "name": order.get("stp_name"),
            "latitude": float(stp_lat),
            "longitude": float(stp_lon)
        },
        "delivery": {
            "latitude": float(delivery_lat),
            "longitude": float(delivery_lon),
            "location": order.get("location", "")
        }
    })


@app.route("/invoice")
def invoice():

    order_id = request.args.get("order_id")

    if not order_id:
        return "Order ID is missing", 400

    order = None

    if os.path.exists(ORDERS_FILE):

        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:

            reader = csv.DictReader(f)

            for row in reader:

                if row.get("order_id") == order_id:
                    order = row
                    break

    if not order:
        return "Order not found", 404

    # Only allow the logged-in buyer to view their own invoice.
    current_user_id = session.get("user_id")
    current_buyer_name = session.get("buyer_name") or session.get("user_name")
    current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

    authorized = (
        current_user_id and
        order.get("buyer_user_id", "") == current_user_id
    ) or (
        not order.get("buyer_user_id", "") and
        current_buyer_name and
        current_buyer_phone and
        order.get("buyer_name") == current_buyer_name and
        order.get("buyer_phone") == current_buyer_phone
    )

    if not authorized:
        return "Unauthorized", 403

    # Convert the existing order data
    # into the names expected by invoice.html

    # =========================================================
    # INVOICE CALCULATION
    # =========================================================

    quantity = float(order.get("quantity_kld") or 0)

    # Price of treated wastewater per KL
    WATER_RATE = 30.0

    # Transportation charge per KL
    TRANSPORT_RATE = 10.0

    # GST rate
    GST_RATE = 0.18

    # Calculate water amount
    water_amount = quantity * WATER_RATE

    # Calculate transportation amount
    transport_amount = quantity * TRANSPORT_RATE

    # Calculate subtotal
    subtotal = water_amount + transport_amount

    # Calculate GST
    gst = subtotal * GST_RATE

    # Calculate final amount
    total = subtotal + gst

    info = {
        "order_id": order.get("order_id"),
        "stp_id": order.get("stp_id"),
        "stp_name": order.get("stp_name"),

        "quantity": order.get("quantity_kld"),
        "quality_required": order.get("quality"),
        "water_type": order.get("water_type"),

        "distance_km": order.get("distance_km"),
        "location": order.get("location"),

        "buyer_name": order.get("buyer_name"),
        "buyer_phone": order.get("buyer_phone"),

        "status": order.get("status"),
        "created_at": order.get("created_at"),

        # Invoice amounts
        "water_rate": f"{WATER_RATE:.2f}",
        "water_amount": f"{water_amount:.2f}",
        "transport_rate": f"{TRANSPORT_RATE:.2f}",
        "transport_amount": f"{transport_amount:.2f}",
        "subtotal": f"{subtotal:.2f}",
        "gst": f"{gst:.2f}",
        "total": f"{total:.2f}"
    }

    return render_template(
        "invoice.html",
        info=info,
        invoice_date=order.get("created_at")
    )


# =========================================================
# PAYMENT / BOOKING CONFIRMATION
# =========================================================

@app.route("/pay_now", methods=["POST"])
def pay_now():
    order_id = request.form.get("order_id", "").strip()

    if not order_id:
        return "Order ID is missing", 400

    if not os.path.exists(ORDERS_FILE):
        return "Orders file not found", 404

    updated_rows = []
    order_found = False

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:
            if row.get("order_id", "").strip() == order_id:
                current_user_id = session.get("user_id")
                current_buyer_name = session.get("buyer_name") or session.get("user_name")
                current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

                authorized = (
                    current_user_id and
                    row.get("buyer_user_id", "") == current_user_id
                ) or (
                    not row.get("buyer_user_id", "") and
                    current_buyer_name and
                    current_buyer_phone and
                    row.get("buyer_name") == current_buyer_name and
                    row.get("buyer_phone") == current_buyer_phone
                )

                if not authorized:
                    return "Unauthorized", 403

                row["status"] = "Pending"
                row["payment_status"] = "Paid"
                order_found = True

            updated_rows.append(row)

    if not order_found:
        return "Order not found", 404

    # Keep every existing column and add payment_status when needed.
    if "payment_status" not in fieldnames:
        fieldnames.append("payment_status")

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)

    return redirect(url_for("demand", payment_success=order_id))


@app.route("/confirm_cod", methods=["POST"])
def confirm_cod():
    order_id = request.form.get("order_id", "").strip()

    if not order_id:
        return "Order ID is missing", 400

    if not os.path.exists(ORDERS_FILE):
        return "Orders file not found", 404

    updated_rows = []
    order_found = False

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:
            if row.get("order_id", "").strip() == order_id:
                current_user_id = session.get("user_id")
                current_buyer_name = session.get("buyer_name") or session.get("user_name")
                current_buyer_phone = session.get("buyer_phone") or session.get("user_phone")

                authorized = (
                    current_user_id and
                    row.get("buyer_user_id", "") == current_user_id
                ) or (
                    not row.get("buyer_user_id", "") and
                    current_buyer_name and
                    current_buyer_phone and
                    row.get("buyer_name") == current_buyer_name and
                    row.get("buyer_phone") == current_buyer_phone
                )

                if not authorized:
                    return "Unauthorized", 403

                row["status"] = "Pending"
                row["payment_status"] = "Cash on Delivery"
                order_found = True

            updated_rows.append(row)

    if not order_found:
        return "Order not found", 404

    # Keep every existing column and add payment_status when needed.
    if "payment_status" not in fieldnames:
        fieldnames.append("payment_status")

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)

    return redirect(url_for("demand", payment_success=order_id))




@app.route("/api/stp_orders")
def stp_orders():
    """Return orders assigned to the logged-in STP operator."""
    if not session.get("user_id"):
        return jsonify({"error": "Please log in to view STP orders."}), 401

    if str(session.get("role", "")).lower().strip() != "stp":
        return jsonify({"error": "Unauthorized"}), 403

    selected_stp_id = str(
        request.args.get("stp_id")
        or session.get("selected_stp_id")
        or ""
    ).strip()

    results = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                row_stp_id = str(row.get("stp_id") or "").strip()

                if selected_stp_id and row_stp_id != selected_stp_id:
                    continue

                results.append({
                    "order_id": row.get("order_id"),
                    "status": row.get("status"),
                    "location": row.get("location"),
                    "stp_id": row.get("stp_id"),
                    "stp_name": row.get("stp_name"),
                    "quantity_kld": row.get("quantity_kld"),
                    "quality": row.get("quality"),
                    "water_type": row.get("water_type"),
                    "distance_km": row.get("distance_km"),
                    "buyer_name": row.get("buyer_name"),
                    "buyer_phone": row.get("buyer_phone"),
                    "created_at": row.get("created_at"),
                    "payment_status": row.get("payment_status", ""),
                    "delivery_lat": row.get("delivery_lat", ""),
                    "delivery_lon": row.get("delivery_lon", "")
                })

    results.sort(key=lambda x: x.get("created_at") or "", reverse=True)

    return jsonify({
        "stp_id": selected_stp_id,
        "orders": results
    })


@app.route("/api/my_orders")
def my_orders():
    user_id = session.get("user_id")
    buyer_name = session.get("buyer_name") or session.get("user_name")
    buyer_phone = session.get("buyer_phone") or session.get("user_phone")

    if not user_id and not buyer_name and not buyer_phone:
        return jsonify({"error": "Please log in to view your orders."}), 401

    results = []
    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                matches_user = bool(user_id and row.get("buyer_user_id", "") == user_id)
                matches_legacy = (
                    not row.get("buyer_user_id", "") and buyer_name and buyer_phone and
                    row.get("buyer_name") == buyer_name and row.get("buyer_phone") == buyer_phone
                )
                if matches_user or matches_legacy:
                    results.append({
                        "order_id": row.get("order_id"),
                        "status": row.get("status"),
                        "location": row.get("location"),
                        "stp_name": row.get("stp_name"),
                        "quantity_kld": row.get("quantity_kld"),
                        "created_at": row.get("created_at"),
                        "payment_status": row.get("payment_status", "")
                    })

    results.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return jsonify(results)


@app.route("/api/track_order")
def track_order():

    order_id = request.args.get("order_id")
    phone = request.args.get("phone")
    
    if not order_id and not phone:
        return jsonify({"error": "Provide order_id or phone"}), 400

    results = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)

            for row in reader:
                if (
                    (order_id and row.get("order_id") == order_id) or
                    (phone and row.get("buyer_phone") == phone)
                ):
                    results.append({
                        "order_id": row.get("order_id"),
                        "status": row.get("status"),
                        "location": row.get("location"),
                        "stp_name": row.get("stp_name"),
                        "created_at": row.get("created_at")
                    })

    results.sort(key=lambda x: x["order_id"], reverse=True)
    return jsonify(results)

# =========================================================
# SUPPLY SIDE
# =========================================================

@app.route('/supply')
def supply():

    if not session.get("user_id"):
        return redirect(url_for("login"))

    if str(session.get("role", "")).lower() != "stp":
        return "Unauthorized", 403

    auto_reset_capacity()

    stps = load_stps()
    selected_id = request.args.get("stp_id")

    # Remember the STP selected by this operator so STP Order Tracking
    # can show the same STP's orders.
    if selected_id:
        session["selected_stp_id"] = str(selected_id).strip()

    selected_stp_id = str(
        session.get("selected_stp_id") or selected_id or ""
    ).strip()

    selected_stp = None
    prediction = None
    weekly_forecast = None

    if selected_id:
        for stp in stps:
            if str(stp["stp_id"]) == str(selected_id):
                selected_stp = stp

                try:
                    print("STP ID sent to ML:", stp["stp_id"])

                    prediction = predict_next_day(str(stp["stp_id"]))
                    weekly_forecast = predict_week(str(stp["stp_id"]))

                    if prediction is not None:
                        prediction = round(prediction, 2)

                    print("Prediction:", prediction)
                except Exception as e:
                    print("Prediction error:", e)
                    prediction = None

    demands = []

    if selected_stp and os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                
                print("ROW STP:", row.get("stp_id"))
                print("SELECTED STP:", selected_stp["stp_id"])

                if row.get("stp_id", "").strip() == str(selected_stp["stp_id"]).strip():
                    
                    print("MATCHED:", row)

                    # ✅ SAFE CLEANING (handles None keys)
                    clean_row = {}

                    for k, v in row.items():
                        if k is None:
                            continue
                        clean_row[k.strip()] = v

                    row = clean_row
                    
                    print("ROW DATA:", row)
                    mapped_row = {
                        "request_id": row.get("order_id"),
                        "site_name": row.get("location"),
                        "quantity": row.get("quantity_kld"),
                        "quality_required": row.get("quality"),
                        "buyer_name": row.get("buyer_name"),        # ✅ ADD THIS
                        "buyer_phone": row.get("buyer_phone"),      # ✅ ADD THIS
                        "status": (row.get("status") or "").strip(),
                        "created_at": row.get("created_at")
                    }

                    demands.append(mapped_row)

    return render_template(
    "supply.html",
    stps=stps,
    selected_stp=selected_stp,
    demands=demands,
    prediction=prediction,
    weekly_forecast=weekly_forecast
    )

@app.route("/update_capacity", methods=["POST"])
def update_capacity():

    stp_id = request.form["stp_id"]
    new_capacity = float(request.form["available_capacity_mld"])

    stps = load_stps()

    for stp in stps:
        if str(stp["stp_id"]) == str(stp_id):
            stp["available_capacity_mld"] = new_capacity

    save_stps(stps)

    return redirect(url_for("supply", stp_id=stp_id))

@app.route("/upload_quality", methods=["POST"])
def upload_quality():

    stp_id = request.form["stp_id"]
    quality = request.form["quality_grade"]

    stps = load_stps()

    for stp in stps:
        if str(stp["stp_id"]) == str(stp_id):
            stp["quality_grade"] = quality

    save_stps(stps)

    return redirect(url_for("supply", stp_id=stp_id))

@app.route("/handle_request", methods=["POST"])
def handle_request():

    if not session.get("user_id"):
        return redirect(url_for("login"))

    if str(session.get("role", "")).lower() != "stp":
        return "Unauthorized", 403

    auto_reset_capacity()


    order_id = request.form["request_id"]
    action = request.form.get("action")

    updated_rows = []
    stp_id_redirect = None


    # STEP 1: READ FILE
    if action not in {"accept", "reject"}:
        return "Invalid action", 400

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or ORDER_FIELDS

        for row in reader:

            if row.get("order_id", "").strip() == order_id.strip():

                stp_id_redirect = row.get("stp_id")
                current_status = (row.get("status") or "").strip()

                # Only pending orders can be accepted/rejected by the STP.
                if current_status != "Pending":
                    updated_rows.append(row)
                    continue

                if action == "reject":
                    row["status"] = "Rejected"
                    updated_rows.append(row)
                    continue

                # =====================================================
                # ACCEPT ORDER ONLY IF IT MATCHES THE STP DATASET
                # =====================================================
                stps = load_stps()
                matching_stp = None

                for stp in stps:
                    if str(stp.get("stp_id")) == str(row.get("stp_id")):
                        matching_stp = stp
                        break

                if matching_stp is None:
                    return "STP not found in STP dataset", 404

                try:
                    quantity_kld = float(row.get("quantity_kld") or 0)
                except (TypeError, ValueError):
                    return "Invalid order quantity", 400

                if quantity_kld <= 0:
                    return "Order quantity must be greater than zero", 400

                quantity_mld = quantity_kld / 1000.0

                try:
                    available_capacity = float(
                        matching_stp.get("available_capacity_mld", 0) or 0
                    )
                except (TypeError, ValueError):
                    available_capacity = 0.0

                # Check available STP capacity.
                if available_capacity < quantity_mld:
                    return "Insufficient STP capacity", 400

                # Check requested quality against the STP dataset.
                requested_quality = (row.get("quality") or "").strip()
                stp_quality = (matching_stp.get("quality_grade") or "").strip()

                if (
                    requested_quality
                    and stp_quality
                    and requested_quality.lower() != stp_quality.lower()
                ):
                    return "Requested water quality is not available at this STP", 400

                # Check requested water type against the STP dataset when
                # the STP has a water_type field populated.
                requested_type = (row.get("water_type") or "").strip()
                stp_type = (matching_stp.get("water_type") or "").strip()

                if (
                    requested_type
                    and stp_type
                    and requested_type.lower() != stp_type.lower()
                ):
                    return "Requested water type is not supported by this STP", 400

                # Reserve the requested quantity.
                matching_stp["available_capacity_mld"] = (
                    available_capacity - quantity_mld
                )

                matching_stp["current_load_mld"] = (
                    float(matching_stp.get("current_load_mld", 0) or 0)
                    + quantity_mld
                )

                accepted_at = datetime.now()
                release_at = accepted_at + timedelta(hours=24)

                row["status"] = "Accepted"
                row["accepted_at"] = accepted_at.isoformat()
                row["capacity_release_at"] = release_at.isoformat()
                row["capacity_released"] = "False"

                save_stps(stps)

            updated_rows.append(row)

    if stp_id_redirect is None:
        return "Order not found", 404

    # Keep every existing order column and the new capacity timeline fields.
    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()
        for row in updated_rows:
            writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})

    return redirect(url_for("supply", stp_id=stp_id_redirect))

@app.route("/update_order_status", methods=["POST"])
def update_order_status():
    
    if not session.get("user_id"):
        return jsonify({"success": False, "error": "Login required"}), 401

    user_role = str(session.get("role", "")).lower().strip()

    auto_reset_capacity()

    order_id = (request.form.get("order_id") or "").strip()
    new_status = (request.form.get("status") or "").strip()

    # STP operators can update the normal workflow.
    # Demand users can only persist Delivered for their own order,
    # which is required when the Track Order tanker animation reaches
    # the exact destination.
    if user_role != "stp":
        if not (
            user_role == "demand"
            and new_status == "Delivered"
        ):
            return jsonify({
                "success": False,
                "error": "Unauthorized"
            }), 403

    allowed_statuses = {"Pending", "Accepted", "Out for Delivery", "Delivered", "Rejected"}
    if new_status not in allowed_statuses:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    if not order_id:
        return jsonify({"success": False, "error": "Order ID is required"}), 400

    updated = False
    updated_rows = []
    stp_id_redirect = None

    with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("order_id", "").strip() == order_id:

                if user_role == "demand":
                    session_user_id = str(
                        session.get("user_id") or ""
                    ).strip()
                    order_user_id = str(
                        row.get("buyer_user_id") or ""
                    ).strip()

                    if (
                        not session_user_id
                        or order_user_id != session_user_id
                    ):
                        return jsonify({
                            "success": False,
                            "error": "Unauthorized"
                        }), 403

                stp_id_redirect = row.get("stp_id")
                current_status = row.get("status", "").strip()
                status_order = {"Pending": 0, "Accepted": 1, "Out for Delivery": 2, "Delivered": 3, "Rejected": -1}

                if (
                    current_status != "Rejected" and new_status != "Rejected" and
                    status_order.get(new_status, -1) < status_order.get(current_status, -1)
                ):
                    return jsonify({"success": False, "error": "Cannot move order backwards"}), 400

                if new_status == "Accepted" and current_status != "Accepted":
                    stps = load_stps()
                    quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
                    stp_found = False
                    for stp in stps:
                        if str(stp.get("stp_id")) == str(row.get("stp_id")):
                            available = float(stp.get("available_capacity_mld", 0) or 0)
                            if quantity_mld > available:
                                return jsonify({"success": False, "error": "Insufficient STP capacity"}), 400
                            stp["available_capacity_mld"] = max(0.0, available - quantity_mld)
                            stp["current_load_mld"] = float(stp.get("current_load_mld", 0) or 0) + quantity_mld
                            stp_found = True
                            break
                    if not stp_found:
                        return jsonify({"success": False, "error": "STP not found"}), 404
                    save_stps(stps)
                    accepted_at = datetime.now()
                    row["accepted_at"] = accepted_at.isoformat()
                    row["capacity_release_at"] = (accepted_at + timedelta(hours=24)).isoformat()
                    row["capacity_released"] = "False"

                if (
                    new_status == "Rejected" and
                    current_status in {"Accepted", "Out for Delivery"} and
                    str(row.get("capacity_released", "")).strip().lower() != "true"
                ):
                    stps = load_stps()
                    quantity_mld = float(row.get("quantity_kld") or 0) / 1000.0
                    for stp in stps:
                        if str(stp.get("stp_id")) == str(row.get("stp_id")):
                            total = float(stp.get("total_capacity_mld") or 0)
                            available = float(stp.get("available_capacity_mld", 0) or 0)
                            stp["available_capacity_mld"] = min(total, available + quantity_mld)
                            stp["current_load_mld"] = max(0.0, float(stp.get("current_load_mld", 0) or 0) - quantity_mld)
                            row["capacity_released"] = "True"
                            save_stps(stps)
                            break

                row["status"] = new_status
                updated = True
            updated_rows.append(row)

    if not updated:
        return jsonify({"success": False, "error": "Order not found"}), 404

    with open(ORDERS_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ORDER_FIELDS)
        writer.writeheader()
        for row in updated_rows:
            writer.writerow({field: row.get(field, "") for field in ORDER_FIELDS})

    if user_role == "demand":
        return jsonify({
            "success": True,
            "order_id": order_id,
            "status": "Delivered"
        })

    return redirect(url_for("supply", stp_id=stp_id_redirect))

@app.route("/stp_track")
def stp_track():
    if not session.get("user_id"):
        return redirect(url_for("login"))

    role = str(session.get("role", "")).lower().strip()

    if role != "stp":
        return jsonify({"error": "Unauthorized"}), 403

    return render_template("stp_track.html")

@app.route("/tanker")
def tanker_dashboard():

    if not session.get("user_id"):
        return redirect(url_for("login"))

    if str(session.get("role", "")).lower() != "tanker":
        return "Unauthorized", 403

    auto_reset_capacity()

    orders = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r") as f:
            reader = csv.DictReader(f)

            for row in reader:

                if row["status"] in {"Accepted", "Out for Delivery"}:

                    stps = load_stps()
                    stp_lat = None
                    stp_lon = None

                    for stp in stps:
                        if str(stp["stp_id"]) == str(row["stp_id"]):
                            stp_lat = stp.get("latitude")
                            stp_lon = stp.get("longitude")
                            break

                    row["stp_lat"] = stp_lat
                    row["stp_lon"] = stp_lon

                    try:
                        row["delivery_lat"] = float(row.get("delivery_lat"))
                        row["delivery_lon"] = float(row.get("delivery_lon"))
                    except (TypeError, ValueError):
                        row["delivery_lat"] = None
                        row["delivery_lon"] = None

                    # Existing tanker.html expects this field.
                    current_order_status = str(row.get("status") or "").strip()
                    row["tanker_request_status"] = (
                        "Accepted"
                        if current_order_status in {"Accepted", "Out for Delivery"}
                        else current_order_status
                    )

                    orders.append(row)

    # Keep the existing active orders list unchanged.
    # Add a separate history list for the tanker page's My Trips section.
    trip_history = []

    if os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)

            for row in reader:
                status = str(row.get("status") or "").strip()

                # My Trips follows the current order status stored in
                # orders.csv. Once an order has entered the tanker workflow,
                # keep the same order in history so its status can move from
                # Accepted -> Out for Delivery -> Delivered.
                if status in {"", "Pending", "Rejected"}:
                    continue

                trip_history.append({
                    "order_id": row.get("order_id", ""),
                    "stp_id": row.get("stp_id", ""),
                    "stp_name": row.get("stp_name", ""),
                    "quantity_kld": row.get("quantity_kld", ""),
                    "quality": row.get("quality", ""),
                    "water_type": row.get("water_type", ""),
                    "location": row.get("location", ""),
                    "buyer_name": row.get("buyer_name", ""),
                    "buyer_phone": row.get("buyer_phone", ""),
                    "distance_km": row.get("distance_km", ""),
                    "status": status,
                    "created_at": row.get("created_at", ""),
                    "delivered_at": row.get("delivered_at", ""),
                    "payment_status": row.get("payment_status", "")
                })

    trip_history.sort(
        key=lambda item: item.get("created_at") or "",
        reverse=True
    )

    return render_template(
        "tanker.html",
        orders=orders,
        trip_history=trip_history
    )

TANKER_CAPACITY_KLD = 12
AVAILABLE_TANKERS = 5

@app.route("/accept_pickup", methods=["POST"])
def accept_pickup():

    if not session.get("user_id"):
        return redirect(url_for("login"))

    if str(session.get("role", "")).lower() != "tanker":
        return "Unauthorized", 403

    order_id = request.form.get("order_id")

    if not order_id:
        return "No Order ID received"

    updated_rows = []
    tanker_info = None
    stp_id_redirect = None

    with open(ORDERS_FILE, "r", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames

        for row in reader:

            # Strip spaces to avoid mismatch
            if row["order_id"].strip() == order_id.strip():
                stp_id_redirect = row["stp_id"]

                if str(row.get("status") or "").strip().lower() != "accepted":
                    updated_rows.append(row)
                    continue

                quantity = float(row["quantity_kld"])

                tankers_required = math.ceil(quantity / TANKER_CAPACITY_KLD)

                tanker_info = {
                    "order_id": row["order_id"],
                    "quantity": quantity,
                    "tankers_required": tankers_required,
                    "available_tankers": AVAILABLE_TANKERS,
                    "sufficient": tankers_required <= AVAILABLE_TANKERS,
                    "buyer_name": row.get("buyer_name"),
                    "buyer_phone": row.get("buyer_phone"),
                }

                row["status"] = "Out for Delivery"

            updated_rows.append(row)

    if tanker_info is None:
        return f"Order {order_id} is not available for pickup. It must be Accepted by the STP first.", 400

    with open(ORDERS_FILE, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(updated_rows)
    
    
    return render_template(
    "tanker_summary.html",
    info=tanker_info,
    stp_id=stp_id_redirect
)

import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))

    import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))

    app.run(
        host="0.0.0.0",
        port=port,
        threaded=True
    )
